import os
import json   # For settings persistence
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import yaml
from ttkthemes import ThemedTk  # For theming  - pip install ttkthemes
import chardet
from pathlib import Path
import re  # <-- Add this line
from datetime import datetime   # <-- Add here, with the other imports
import threading
import numpy as np
import logging  # Add this import
import traceback
import queue  # For thread-safe logging
from openpyxl import load_workbook  # For Excel manipulation
import shutil  # For file backups
from tkinter import messagebox
from openpyxl.styles import Font, Alignment



def categorize_domain(email):
    if pd.isna(email) or '@' not in str(email):
        return 'other'
    domain = email.split('@')[-1].lower()
    if domain == 'gmail.com': return 'gmail'
    if domain in ['yahoo.com', 'aol.com']: return 'yahoo/aol'
    return 'other'

CONFIG_FILE = 'user_settings.json'

# --- Tooltip Helper ---
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw, text=self.text, justify=tk.LEFT,
            background="#ffffe0", relief=tk.SOLID, borderwidth=1,
            font=("TkDefaultFont", 10)
        )
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()


class ReportTracker:
    def __init__(self):
        self.initialize_metrics()

    def initialize_metrics(self):
        self.metrics = {
            'stepwise': [],  # List of dicts for each step
            'initial_count': 0,
            'invalid_sales_type': 0,          # "Sales type"
            'invalid_model_year': 0,          # "Invalid Model Year"
            'invalid_emails': 0,
            'blank_zip': 0,
            'blank_name': 0,
            'duplicate_addresses': 0,
            'bad_state_codes': 0,             # "Bad State Codes" (split from 'bad_geography')
            'bad_zip_state': 0,               # "Bad State/ZIP" (split from 'bad_geography')
            'business_exclusions': 0,         # "Businesses"
            'duplicate_vins': 0,              # "Duplicate VINS" (NEW METRIC)
            'vin_rejected': 0,
            'vin_missing_cell': 0,
            'vin_missing_modelid': 0,
            'dedup_initial': 0,
            'dedup_removed': 0,
            'dedup_remaining': 0,
            'historical_deduped': 0,
            'truecar_deduped': 0,
            'missing_descriptions': 0,
            'blank_sales_type': 0,
            'sales_types': {},                 # Separate dict for sales type counts
            'final_panels': 0,                # Added for final metrics
            'final_panel_records': 0           # Added for final metrics
        }

    def log_step(self, step_name, before, rejected, after):
        # Call this after each filter step
        self.metrics['stepwise'].append({
            'Step': step_name,
            'Records Before': before,
            'Rejected': rejected,
            'Records After': after
        })

    def log_final_metrics(self, panel_counts):
        """Log final pipeline metrics"""
        total_panels = len(panel_counts)
        total_records = sum(panel_counts.values())
        self.metrics['final_panels'] = total_panels
        self.metrics['final_panel_records'] = total_records

    def update(self, key, value):
        if key == 'sales_type':
            # For sales types, value should be a tuple: (type_name, count)
            stype, count = value
            self.metrics['sales_types'][stype] = self.metrics['sales_types'].get(stype, 0) + count
        else:
            self.metrics[key] = self.metrics.get(key, 0) + value

    def set(self, key, value):
        self.metrics[key] = value

    def get(self, key, default=None):
        return self.metrics.get(key, default)

    def get_metrics(self):
        return self.metrics.copy()

    def reset(self):
        self.initialize_metrics()


import threading

# --- Tooltip Helper ---
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 25
        y = y + cy + self.widget.winfo_rooty() + 25
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            tw, text=self.text, justify=tk.LEFT,
            background="#ffffe0", relief=tk.SOLID, borderwidth=1,
            font=("TkDefaultFont", 10)
        )
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class DataProcessorApp:
    CONFIG_FILE = 'user_settings.json'
    VIN_COLUMN_NAMES = ['vin', 'vehicleid', 'usa_vin']

    def __init__(self, root):
        print("DEBUG: Entered DataProcessorApp.__init__")
        self.root = root
        self.root.title("Universal Vehicle Data Processor")
        self.root.geometry("1200x800")
        self.root.minsize(900, 600)
        self.tracker = ReportTracker()
        self.email_pattern = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        self.progress_queue = queue.Queue()

        # Configuration paths
        self.base_dir = Path(__file__).resolve().parent
        self.config_dir = self.base_dir / "config"
        self.vin_mapping_dir = self.config_dir / "vin_mappings"

        # Initialize file lists to empty so they're always defined
        self.historical_files = []
        self.truecar_files = []

        # Load last-used directories/settings
        self.last_dirs = self.load_settings()

        self.create_menu()
        self.create_widgets()
        self.set_theme('aquativo')  # or 'scidblue', 'scidgreen', etc.
        self.status("Ready.")

        print("DEBUG: Finished DataProcessorApp.__init__")

    def standardize_columns(self, df):
        """Ensure consistent column naming across pipeline"""
        column_aliases = {
            # Existing standardizations
            'ZIP_CODE': 'ZIP',     # For uppercase variant
            'zip_code': 'ZIP',     # For lowercase variant
            'CELL': 'SVI CELL',
            'MODEL_YEAR': 'MOD YR',
            'MODYY': 'MOD YR',
            'PURCH_LSE': 'PURCH/LSE',
            'SEQUENCE': 'SEQUENCE #(8)',
            
            # New bodycode mappings
            'model_id': 'bodycode',          # Acura/Honda
            'NA_MODEL_CODE': 'bodycode',     # BMW/MINI
            'ModelCode': 'bodycode',         # Infiniti/Nissan
            'MODEL_DESC': 'bodycode',        # Mercedes
            'ProductionCode': 'bodycode',    # Volvo (if exists)
            'ModelDescription': 'bodycode',  # Volvo alternative
        }
        
        # Rename columns using aliases
        for alias, standard in column_aliases.items():
            if alias in df.columns:
                df.rename(columns={alias: standard}, inplace=True)
        
        return df

    def update_progress(self):
        try:
            while True:
                value = self.progress_queue.get(block=False)
                self.progress['value'] = value
                self.status_var.set(f"Progress: {value:.1f}%")
        except queue.Empty:
            pass
        self.root.after(100, self.update_progress)


    def _get_vin_mapping_path(self, manufacturer, vin_mapping_dir):
        """Find the correct VIN mapping YAML file for the manufacturer."""
        manufacturer = manufacturer.lower().strip()
        for file in vin_mapping_dir.glob("*.yaml"):
            if file.is_file() and file.stem == manufacturer:
                return file
        raise FileNotFoundError(f"No VIN mapping YAML file found for manufacturer: {manufacturer}")


    def _get_column_mapping_path(self, manufacturer):
        """Find the correct column mapping YAML file for the manufacturer."""
        manufacturer = manufacturer.lower().strip()
        colmap_dir = self.config_dir / "column_mapping"
        for file in colmap_dir.glob("*.yaml"):
            # Accepts both acura_column_map.yaml and honda_column_map.yaml, etc.
            if file.is_file() and manufacturer in file.stem:
                return file
        raise FileNotFoundError(f"No column mapping YAML file found for manufacturer: {manufacturer}")

    @staticmethod
    def remove_decimal_zero(series):
        return series.str.replace(r'\.0$', '', regex=True)

    def set_theme(self, theme_name):
        style = ttk.Style(self.root)
        try:
            style.theme_use(theme_name)
        except tk.TclError:
            style.theme_use('clam')  # fallback

    def create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Save Settings", command=self.save_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help", command=self.show_help)
        help_menu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)
        self.root.config(menu=menubar)

    def show_help(self):
        messagebox.showinfo("Help", "Instructions for using the Universal Vehicle Data Processor:\n\n"
            "- Fill in all required fields.\n"
            "- Use the buttons to select files and directories.\n"
            "- Click 'Start Processing' to begin.\n"
            "- The log and status bar will show progress and errors.\n"
            "- For more details, see the documentation or contact support.")

    def show_about(self):
        messagebox.showinfo("About", "Universal Vehicle Data Processor\nVersion 1.0\n© 2025 Your Company")

    def get_manufacturers(self):
        yaml_files = [f.stem for f in self.vin_mapping_dir.glob("*.yaml")]
        return sorted(yaml_files)

    def create_widgets(self):
        print("DEBUG: Entered create_widgets()")

        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.grid(row=0, column=0, sticky="nsew")

        # --- Title Label (NEW) ---
        title_label = ttk.Label(
            main_frame,
            text="Universal Vehicle Data Processor",
            font=("Segoe UI", 16, "bold")
        )
        title_label.grid(row=0, column=0, pady=(0, 15))

        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        # ============================
        # 1. Input Parameters (row 1)
        # ============================
        input_frame = ttk.LabelFrame(main_frame, text="Input Parameters", padding=10)
        input_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=10)
        for i in range(8):
            input_frame.columnconfigure(i, weight=1)

        ttk.Label(input_frame, text="Manufacturer:").grid(row=0, column=0, sticky='e')
        self.manufacturer_combo = ttk.Combobox(input_frame, values=self.get_manufacturers(), width=20, state='readonly')
        self.manufacturer_combo.grid(row=0, column=1, sticky='w')
        ToolTip(self.manufacturer_combo, "Select the OEM manufacturer.")

        ttk.Label(input_frame, text="Month:").grid(row=0, column=2, sticky='e')
        self.month_entry = ttk.Entry(input_frame, width=10)
        self.month_entry.grid(row=0, column=3, sticky='w')
        ToolTip(self.month_entry, "Enter the processing month (e.g., 06).")

        ttk.Label(input_frame, text="Year:").grid(row=0, column=4, sticky='e')
        self.year_entry = ttk.Entry(input_frame, width=10)
        self.year_entry.grid(row=0, column=5, sticky='w')
        ToolTip(self.year_entry, "Enter the processing year (e.g., 2025).")

        ttk.Label(input_frame, text="Sequence #:").grid(row=0, column=6, sticky='e')
        self.sequence_entry = ttk.Entry(input_frame, width=20)
        self.sequence_entry.grid(row=0, column=7, sticky='w')
        ToolTip(self.sequence_entry, "Enter the starting sequence number.")

        # ============================
        # 2. File Selection (row 2)
        # ============================
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding=10)
        file_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=10)
        for i in range(9):
            file_frame.columnconfigure(i, weight=1)

        ttk.Button(file_frame, text="Select Input", command=self.select_input).grid(row=0, column=0, sticky='e')
        self.input_entry = ttk.Entry(file_frame, width=100)
        self.input_entry.grid(row=0, column=1, columnspan=7, sticky='ew')
        ToolTip(self.input_entry, "Path to the input data file.")

        ttk.Label(file_frame, text="Input file should be a CSV saved as UTF-8").grid(row=0, column=8, sticky='w')

        ttk.Button(file_frame, text="Select Output", command=self.select_output).grid(row=1, column=0, sticky='e')
        self.output_entry = ttk.Entry(file_frame, width=100)
        self.output_entry.grid(row=1, column=1, columnspan=7, sticky='ew')
        ToolTip(self.output_entry, "Directory for all output files.")

        ttk.Label(file_frame, text="UCC Master:").grid(row=2, column=0, sticky='e')
        self.ucc_entry = ttk.Entry(file_frame, width=100)
        self.ucc_entry.grid(row=2, column=1, columnspan=7, sticky='ew')
        ttk.Button(file_frame, text="Browse", command=self.select_ucc_file).grid(row=2, column=8, sticky='w')

        ttk.Label(file_frame, text="Electric Vehicles File:").grid(row=3, column=0, sticky='e')
        self.electric_entry = ttk.Entry(file_frame, width=100)
        self.electric_entry.grid(row=3, column=1, columnspan=7, sticky='ew')
        ttk.Button(file_frame, text="Browse", command=lambda: self.select_file(self.electric_entry)).grid(row=3, column=8, sticky='w')

        ttk.Label(file_frame, text="Description File:").grid(row=4, column=0, sticky='e')
        self.desc_entry = ttk.Entry(file_frame, width=100)
        self.desc_entry.grid(row=4, column=1, columnspan=7, sticky='ew')
        ttk.Button(file_frame, text="Browse", command=lambda: self.select_file(self.desc_entry)).grid(row=4, column=8, sticky='w')

        # ============================
        # 3. Deduplication Sources (row 3)
        # ============================
        dedupe_frame = ttk.LabelFrame(main_frame, text="Deduplication Sources", padding=10)
        dedupe_frame.grid(row=3, column=0, sticky="ew", padx=5, pady=10)
        for i in range(4):
            dedupe_frame.columnconfigure(i, weight=1)

        ttk.Button(dedupe_frame, text="Add Files", command=self.add_historical_files).grid(row=0, column=2, sticky='w')
        ttk.Button(dedupe_frame, text="Clear", command=lambda: self.hist_listbox.delete(0, tk.END)).grid(row=0, column=3, sticky='w')

        ttk.Label(dedupe_frame, text="Historical Files:").grid(row=0, column=0, sticky='e')
        self.hist_listbox = tk.Listbox(dedupe_frame, height=3, width=80)
        self.hist_listbox.grid(row=0, column=1, sticky='ew')

        ttk.Label(dedupe_frame, text="TrueCar Files:").grid(row=1, column=0, sticky='e')
        self.truecar_listbox = tk.Listbox(dedupe_frame, height=5, width=80)
        self.truecar_listbox.grid(row=1, column=1, sticky='ew')
        self.truecar_scrollbar = ttk.Scrollbar(dedupe_frame, orient=tk.VERTICAL, command=self.truecar_listbox.yview)
        self.truecar_scrollbar.grid(row=1, column=2, sticky='ns')
        self.truecar_listbox.config(yscrollcommand=self.truecar_scrollbar.set)
        ttk.Button(dedupe_frame, text="Add Files", command=self.add_truecar_files).grid(row=1, column=3, sticky='w')

        # ============================
        # 4. Processing Controls (row 4)
        # ============================
        control_frame = ttk.LabelFrame(main_frame, text="Processing", padding=10)
        control_frame.grid(row=4, column=0, sticky="ew", padx=5, pady=10)
        control_frame.columnconfigure(0, weight=1)

        self.progress = ttk.Progressbar(control_frame, mode='determinate')
        self.progress.grid(row=0, column=0, sticky="ew", padx=2, pady=2)

        self.btn_process = ttk.Button(control_frame, text="Start Processing", command=self.start_processing)
        self.btn_process.grid(row=1, column=0, pady=5)

        # ============================
        # 5. Logging (row 5)
        # ============================
        log_frame = ttk.LabelFrame(main_frame, text="Processing Log", padding=5)
        log_frame.grid(row=5, column=0, sticky="nsew", padx=5, pady=10)
        main_frame.rowconfigure(5, weight=1)
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(log_frame, wrap=tk.WORD, font=("Consolas", 11), height=20, width=120)
        self.log_text.grid(row=0, column=0, sticky="nsew")

        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        log_scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_text['yscrollcommand'] = log_scrollbar.set

        # ============================
        # 6. Status Bar (root row 1)
        # ============================
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor='w', padding=5)
        status_bar.grid(row=1, column=0, sticky="ew")

        # ============================
        # 7. Summary Button (row 6)
        # ============================
        self.summary_button = ttk.Button(main_frame, text="Generate Summary", command=self.generate_summary_report)
        self.summary_button.grid(row=6, column=0, pady=10)

        print("DEBUG: Finished create_widgets()")

    def status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

    # --- File dialog helpers with last-dir memory ---
    def select_file(self, entry_widget):
        initialdir = self.last_dirs.get('input', os.getcwd())
        path = filedialog.askopenfilename(initialdir=initialdir, filetypes=[("All Files", "*.*")])
        if path:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, path)
            self.last_dirs['input'] = str(Path(path).parent)
            self.save_settings()
            self.status(f"Selected file: {path}")

    def select_input(self):
        initialdir = self.last_dirs.get('input', os.getcwd())
        path = filedialog.askopenfilename(
            title="Select Input File",
            initialdir=initialdir,
            filetypes=[("PRN/CSV/Excel", "*.prn *.csv *.xls *.xlsx"), ("All Files", "*.*")]
        )
        if path:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, path)
            self.last_dirs['input'] = str(Path(path).parent)
            self.save_settings()
            self.status(f"Selected input file: {path}")

    def select_output(self):
        initialdir = self.last_dirs.get('output', os.getcwd())
        path = filedialog.askdirectory(title="Select Output Directory", initialdir=initialdir)
        if path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, path)
            self.last_dirs['output'] = path
            self.save_settings()
            self.status(f"Selected output directory: {path}")

    def select_ucc_file(self):
        initialdir = self.last_dirs.get('ucc', os.getcwd())
        path = filedialog.askopenfilename(
            title="Select UCC Master File",
            initialdir=initialdir,
            filetypes=[("PRN/CSV/Excel", "*.prn *.csv *.xls *.xlsx"), ("All Files", "*.*")]
        )
        if path:
            self.ucc_entry.delete(0, tk.END)
            self.ucc_entry.insert(0, path)
            self.last_dirs['ucc'] = str(Path(path).parent)
            self.save_settings()
            self.status(f"Selected UCC master file: {path}")

    def add_historical_files(self):
        files = filedialog.askopenfilenames(filetypes=[("CSV/Excel", "*.csv *.xlsx")])
        for path in files:
            if Path(path).exists():
                self.hist_listbox.insert(tk.END, path)  # Store full path in listbox
            else:
                messagebox.showwarning("File Not Found", f"File not found:\n{path}")


    def add_truecar_files(self):
        files = filedialog.askopenfilenames(
            title="Select TrueCar Files",
            filetypes=[("CSV/Excel", "*.csv *.xlsx")]
        )
        self.truecar_listbox.delete(0, tk.END)
        for path in files:
            if Path(path).exists():
                self.truecar_listbox.insert(tk.END, path)  # Store full path
            else:
                messagebox.showwarning("File Not Found", f"File not found:\n{path}")


    def log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        self.status(message)

    def start_processing(self):
        print("DEBUG: Start processing button clicked")
        self.log("Start processing button clicked")
        if not self.validate_inputs():
            self.log("Validation failed")
            return
        self.btn_process["state"] = "disabled"
        self.progress["value"] = 0
        self.tracker.reset()
        self.log_queue = queue.Queue()
        self.log("Starting processing thread")
        processing_thread = threading.Thread(target=self.run_pipeline, daemon=True)
        processing_thread.start()
        self.process_logs()
        self.update_progress()  # Add this line
        self.check_thread(processing_thread)
        print("DEBUG: Processing thread started")


    def process_logs(self):
        try:
            while True:
                message = self.log_queue.get(block=False)  # Nowait equivalent
                self.log_text.insert(tk.END, f"{message}\n")
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        self.root.after(100, self.process_logs)

    def thread_log(self, message):
        self.log_queue.put(message)

    def check_thread(self, thread):
        if thread.is_alive():
            self.root.after(100, lambda: self.check_thread(thread))
        else:
            self.btn_process["state"] = "normal"
            self.progress["value"] = 0

    def validate_inputs(self):
        # Get file lists from listboxes
        historical_files = self.hist_listbox.get(0, tk.END)
        truecar_files = self.truecar_listbox.get(0, tk.END)
        
        checks = [
            (self.input_entry.get(), "Input file"),
            (self.output_entry.get(), "Output directory"),
            (self.manufacturer_combo.get(), "Manufacturer"),
            (historical_files, "Historical files"),  # Now uses listbox
            (truecar_files, "TrueCar files"),        # Now uses listbox
            (self.ucc_entry.get(), "UCC master file")
        ]
        
        # Validate non-empty fields
        for val, name in checks:
            if not val or (isinstance(val, tuple) and len(val) == 0):
                self.log(f"Validation failed: {name} is missing or empty.")
                messagebox.showwarning("Missing Inputs", f"{name} is required")
                return False
        
        # Validate file existence
        required_paths = [
            (self.input_entry.get(), "Input file"),
            (self.ucc_entry.get(), "UCC master file")
        ]
        
        for path, name in required_paths:
            if not Path(path).exists():
                self.log(f"Validation failed: {name} not found at {path}")
                messagebox.showerror("Missing File", f"{name} not found: {path}")
                return False
                
        # Validate output directory
        output_dir = Path(self.output_entry.get())
        if not output_dir.exists() or not output_dir.is_dir():
            self.log(f"Validation failed: Output directory invalid: {output_dir}")
            messagebox.showerror("Invalid Directory", "Output directory must be an existing folder")
            return False
            
        return True


    def load_settings(self):
        if os.path.exists(self.CONFIG_FILE):
            with open(self.CONFIG_FILE, 'r') as f:
                return json.load(f)
        return {}

    def save_settings(self):
        with open(self.CONFIG_FILE, 'w') as f:
            json.dump(self.last_dirs, f)


# Usage example:
# if __name__ == "__main__":
#     root = ThemedTk(theme="aquativo")  # or "scidblue", "scidgreen", etc.
#     app = DataProcessorApp(root)
#     root.mainloop()



    def run_pipeline(self):
        print("DEBUG: Entered run_pipeline")  # Bookend at start
        self.log("Pipeline started")
        steps = [
            self.step_email_validation,         # 1. Email + blank name cleanup
            self.step_deduplicate_addresses,    # 2. Filter duplicate addresses (NEW - add this)
            self.step_sales_type_filter,        # 3. Purch/LSE handling
            self.step_geography_filter,         # 4. State/ZIP filtering
            self.step_business_filter,          # 5. Business exclusion
            self.step_vin_processing,           # 6. VIN mapping + EXTERNAL/REFERENCE
            self.step_deduplication,            # 7. Duplicate removal
            self.step_ucc_check,                # 8. Identify missing UCC combinations
            self.step_ucc_merge,                # 9. Merge UCC codes (FIRST PASS - may fail)
            self.step_electric_merge,           # 10. EV flag
            self.step_desc_merge,               # 11. Add descriptions
            self.step_assign_sequence,          # 12. Sequence numbers
            self.step_combine_rejections,       # 13. Master rejection report
            self.step_cellcode_reporting,       # 14. Cell code report
            self.step_final_outputs,            # 15. FINAL OUTPUT (all columns present)
            self.step_panelization,             # 16. NEW: Email/mail split + domain panels 
            self.step_metrics_reporting         # 17. Metrics reporting
        ]
        try:
            for idx, step in enumerate(steps):
                step_name = step.__name__.replace("step_", "").replace("_", " ").title()
                self.log(f"Processing step {idx+1}/{len(steps)}: {step_name}")
                self.status(f"Step {idx+1}/{len(steps)}: {step_name}")
                if not step():
                    raise RuntimeError(f"Step {idx+1} failed")
                self.progress["value"] = (idx+1)/len(steps)*100

            if hasattr(self, 'ucc_updated') and self.ucc_updated:
                self.log("\nRe-processing with updated UCC master...")
                self.tracker.reset()
                for idx, step in enumerate(steps[5:], start=5):
                    self.log(f"Re-processing step {idx+1}/{len(steps)}")
                    if not step():
                        raise RuntimeError(f"Step {idx+1} failed")
                    self.progress["value"] = (idx+1)/len(steps)*100

            print("DEBUG: Finished run_pipeline successfully")
            messagebox.showinfo("Success", "Processing completed!")
        except Exception as e:
            print(f"DEBUG: Exception in run_pipeline: {e}")
            self.log(f"Critical error: {str(e)}")
            messagebox.showerror("Fatal Error", str(e))
        finally:
            print("DEBUG: Exiting run_pipeline (finally block)")
            self.btn_process["state"] = "normal"



    def step_email_validation(self):
        print("DEBUG: Entered step_email_validation")
        try:
            self.log_queue.put("Starting email validation...")

            output_dir = Path(self.output_entry.get())
            self.log(f"Email output will be saved to: {output_dir}")

            manufacturer = self.manufacturer_combo.get().lower().strip()
            map_path = self.config_dir / "column_mapping" / f"{manufacturer}_column_map.yaml"

            # ===================
            # -- Load DataFrame --
            # ===================
            if manufacturer == "volvo":
                with open(map_path, encoding="utf-8") as f:
                    mapping = yaml.safe_load(f)

                df = pd.read_csv(self.input_entry.get(), header=None, dtype=str)
                df = self.standardize_columns(df)
                self.df = df
                df.columns = [mapping['columns'].get(i, f"col{i}") for i in range(df.shape[1])]
                for col in mapping.get('optional_columns', []):
                    if col not in df.columns:
                        df[col] = ''
                self.log(f"Rows after reading Volvo input: {len(df)}")

            elif manufacturer == "jlr":
                with open(map_path, encoding="utf-8") as f:
                    mapping_yaml = yaml.safe_load(f)

                df = pd.read_csv(self.input_entry.get(), dtype=str, encoding='utf-8')
                raw_headers = [col.strip() for col in df.columns]

                mapping = (
                    mapping_yaml['legacy_mapping']
                    if any("bapibus1006" in col.lower() for col in raw_headers)
                    else mapping_yaml['current_mapping']
                )

                df = self.standardize_columns(df)
                self.df = df
                df.columns = [col.strip() for col in df.columns]

                df.rename(columns=mapping['columns'], inplace=True)

                self.log(f"🔍 Remapped JLR columns: {df.columns.tolist()}")

                # Optional fallback mapping
                if 'modelid' not in df.columns and 'modelcode' in df.columns:
                    df.rename(columns={'modelcode': 'modelid'}, inplace=True)
                    self.log("Fallback: Renamed modelcode to modelid")

                # Handle address merge logic
                df = self.consolidate_address_columns(df, manufacturer)

                for col in mapping_yaml.get('optional_columns', []):
                    if col not in df.columns:
                        df[col] = ''
                
                self.log(f"Rows after reading JLR input: {len(df)}")


                df = self.standardize_columns(df)
                self.df = df
                df.columns = [col.strip() for col in df.columns]
                df.rename(columns=mapping['columns'], inplace=True)
                if 'modelid' not in df.columns and 'bodycode' in df.columns:
                    df['modelid'] = df['bodycode']
                    self.log("modelid populated from bodycode (fallback)")
                df = self.consolidate_address_columns(df, manufacturer)
                for col in mapping.get('optional_columns', []):
                    if col not in df.columns:
                        df[col] = ''
                self.log(f"Rows after reading JLR input: {len(df)}")
            else:
                with open(map_path, encoding="utf-8") as f:
                    mapping = yaml.safe_load(f)
                try:
                    df = pd.read_csv(self.input_entry.get(), dtype=str, encoding='utf-8')
                    self.log("Loaded input file with UTF-8 encoding.")
                except UnicodeDecodeError:
                    df = pd.read_csv(self.input_entry.get(), dtype=str, encoding='ISO-8859-1')
                    self.log("Loaded input file with ISO-8859-1 encoding.")
                df = self.standardize_columns(df)
                self.df = df
                df.columns = [col.strip().lower() for col in df.columns]
                df.rename(columns=mapping['columns'], inplace=True)
                df = self.consolidate_address_columns(df, manufacturer)
                # Special cases...
                if manufacturer == "bmw":
                    phone_cols = ['mobile_phone', 'home_phone', 'work_phone']
                    available = [col for col in phone_cols if col in df.columns]
                    df['phone'] = ''
                    for col in available:
                        df['phone'] = df['phone'].fillna(df[col])
                if manufacturer in ["honda", "acura"]:
                    if 'pid_id' in df.columns:
                        if 'extra1' in df.columns:
                            df['extra1'] = df['extra1'].fillna(df['pid_id'])
                        else:
                            df.rename(columns={'pid_id': 'extra1'}, inplace=True, errors='ignore')
                        df.drop('pid_id', axis=1, inplace=True, errors='ignore')
                    self.log("Populated extra1 from pid_id")
                # Add missing columns from YAML
                all_expected_columns = set(mapping['columns'].values())
                optional_columns = set(mapping.get('optional_columns', []))
                expected = all_expected_columns | optional_columns
                for col in expected:
                    if col not in df.columns:
                        df[col] = ''
                        self.log(f"Added missing column from YAML: {col}")

                legacy_zip_brands = ['honda', 'acura', 'infiniti', 'nissan']
                if manufacturer in legacy_zip_brands and 'zip_code' in df.columns:
                    df['zip_code'] = df['zip_code'].astype(str).str.strip()
                    df['zip_code5'] = df['zip_code'].str[:5]
                    df['zip4'] = df['zip_code'].str[6:10]
                    no_hyphen_mask = df['zip_code'].str.len() == 9
                    df.loc[no_hyphen_mask, 'zip_code'] = df['zip_code'].str[:5]
                    df.loc[no_hyphen_mask, 'zip4'] = df['zip_code'].str[5:]

            initial_count = len(df)
            self.log(f"Email Validation: Starting with {initial_count} records")
            self.tracker.set('initial_count', initial_count)

            # =============================
            # -- Safe Name Construction  --
            # =============================
            # If no name, try to build it from first_name + last_name or parts
            build_attempted = False
            if 'name' not in df.columns or df['name'].isna().all() or df['name'].str.strip().eq('').all():
                name_parts_found = [col for col in ['first_name', 'last_name', 'name_prefix', 'middle_name', 'name_suffix'] if col in df.columns]
                self.log(f"Building 'name' using columns: {name_parts_found}")
                if 'first_name' in df.columns and 'last_name' in df.columns:
                    df['name'] = (
                        df['first_name'].fillna('') + ' ' + df['last_name'].fillna('')
                    ).str.replace(r'\s+', ' ', regex=True).str.strip()
                    self.log("Constructed 'name' from first_name + last_name.")
                    build_attempted = True
                elif name_parts_found:
                    df['name'] = df[name_parts_found].fillna('').agg(' '.join, axis=1).str.replace(r'\s+', ' ', regex=True).str.strip()
                    self.log("Constructed 'name' from available name parts.")
                    build_attempted = True
                else:
                    self.log("No columns to build a name—will reject all as blank name.")
                    df['name'] = ''

            self.log(f"Sample names: {df['name'].dropna().head().tolist()}")

            # ========================
            # --- ZIP CODE DETECT  ---
            # ========================
            zip_col = None
            for candidate in ['zipcode', 'zip_code']:
                if candidate in df.columns:
                    zip_col = candidate
                    break

            if zip_col:
                zip_sample = df[zip_col].dropna().astype(str).str.strip().head().tolist()
                self.log(f"First 5 ZIP codes from column '{zip_col}': {zip_sample}")
            else:
                self.log("No ZIP code column found (checked: 'zipcode', 'zip_code').")

            # ================
            # -- BLANK NAME FILTER --
            # ================
            before = len(df)
            blank_name_df = df[df['name'].isna() | (df['name'].str.strip() == '')]
            valid_df = df[~df.index.isin(blank_name_df.index)]
            removed = before - len(valid_df)
            self.log(f"Removed {removed} records with blank names")
            self.tracker.set('blank_name', removed)

            # --- Save removed blank names if any ---
            name_rem_path = output_dir / "name_removed_blanks.csv"
            if not blank_name_df.empty:
                blank_name_df.to_csv(name_rem_path, index=False)
                self.log(f"Saved {len(blank_name_df)} records with blank names to: {name_rem_path}")
            else:
                self.log("No blank-name rows to remove.")

            # --- Blank ZIPs ---
            blank_zip = (valid_df[zip_col] == '').sum() if zip_col else 0
            self.tracker.set('blank_zip', blank_zip)

            addr_cols = ['address_line1', 'city', 'state', 'zip_code']
            if all(c in valid_df.columns for c in addr_cols):
                dup_groups = valid_df[valid_df.duplicated(addr_cols, keep=False)].groupby(addr_cols).size()
                self.tracker.set('duplicate_addresses', len(dup_groups))

            # --- Email Field Mapping ---
            email_col = None
            for cand in ['email', 'email_addr', 'email_address']:
                if cand in valid_df.columns:
                    email_col = cand
                    break
            if not email_col:
                raise ValueError("Missing required email column.")
            if email_col != 'email':
                valid_df.rename(columns={email_col: 'email'}, inplace=True)
            valid_df['email'] = valid_df['email'].str.strip().str.lower()

            # --- Email Rules ---
            with open(self.config_dir / "email_rules.yml", encoding="utf-8") as f:
                rules = yaml.safe_load(f)
            domain_pattern = re.compile(r'@(?:{})'.format('|'.join(rules['full_domain_rejects'])))
            partial_pattern = re.compile(r'\b(?:{})\b'.format('|'.join(rules['partial_matches'])))

            # --- Validate Emails ---
            valid_rows, rejected_rows = [], []
            for idx, row in valid_df.iterrows():
                email = row.get('email', '')
                # ✔️ Allow blank or missing emails to pass validation
                if pd.isna(email) or str(email).strip() == '':
                    valid_rows.append(row.to_dict())
                    continue
                email = str(email).strip().lower()
                if (domain_pattern.search(email) or
                    partial_pattern.search(email) or
                    not re.match(self.email_pattern, email)):
                    row_dict = row.to_dict()
                    row_dict['rejection_reason'] = 'Invalid email'
                    rejected_rows.append(row_dict)
                    continue
                valid_rows.append(row.to_dict())

            valid_email_df = pd.DataFrame(valid_rows)
            rejected_email_df = pd.DataFrame(rejected_rows)

            blank_count = sum(1 for r in valid_rows if not str(r.get('email', '')).strip())
            invalid_count = len(rejected_rows)
            valid_email_count = len(valid_email_df) - blank_count
            total_passed = len(valid_email_df)

            self.log(f"Email validation complete:")
            self.log(f"  ✔ Valid emails: {valid_email_count}")
            self.log(f"  ✔ Blank emails allowed: {blank_count}")
            self.log(f"  ❌ Invalid emails rejected: {invalid_count}")
            self.log(f"  ➞ Total passed (valid + blank): {total_passed}")

            self.progress_queue.put(100)

            # --- Save Email Rejected ---
            email_rejected_path = output_dir / "email_rejected.csv"
            if not rejected_email_df.empty:
                rejected_email_df.to_csv(email_rejected_path, index=False)
                self.log(f"Wrote {len(rejected_email_df)} invalid email records to {email_rejected_path}")
            else:
                self.log("No invalid emails to reject. Skipped writing email_rejected.csv.")

            self.log(f"Email Validation: {len(rejected_email_df)} invalid emails rejected, {len(valid_email_df)} valid remain")
            self.tracker.set('invalid_emails', len(rejected_email_df))

            # --- STOP EARLY IF EVERYTHING WAS REJECTED ---
            if valid_email_df.empty:
                self.log("No records passed name/email validation. Skipping email cleaned output.")
                self.log("Step finished: No valid records remain to proceed to next step.")
                return False

            # --- Save Cleaned Output ---
            email_cleaned_path = output_dir / "email_cleaned.csv"
            if not valid_email_df.empty:
                valid_email_df.to_csv(email_cleaned_path, index=False)
                self.df = valid_email_df
                self.log(f"Wrote {len(valid_email_df)} valid records to {email_cleaned_path}")
                self.status("Email validation completed.")
                print("DEBUG: Finished step_email_validation successfully")
                return True
            else:
                self.log("No valid email/name records. Skipping email_cleaned.csv")
                return False

        except Exception as e:
            print(f"DEBUG: Exception in step_email_validation: {e}")
            self.log_queue.put(f"Email validation failed: {str(e)}")
            self.log(f"Email validation failed: {str(e)}")
            self.status("Email validation failed.")
            self.progress_queue.put(0)
            return False



    def consolidate_address_columns(self, df, manufacturer):
        manufacturer = manufacturer.lower().strip()
        
        if manufacturer == "jlr":
            parts = []
            if 'address_house_no' in df.columns:
                parts.append(df['address_house_no'].fillna(''))
            if 'address_line1' in df.columns:
                parts.append(df['address_line1'].fillna(''))
            if 'address_house_no2' in df.columns:
                parts.append(df['address_house_no2'].fillna(''))
            
            if parts:
                df['address_line1'] = (
                    parts[0] + ' ' + parts[1] + ' ' + parts[2] if len(parts) == 3 else
                    ' '.join(parts)
                ).str.replace(r'\s+', ' ', regex=True).str.strip()
                self.log("JLR address_line1 constructed from house_no, line1, house_no2")
            
            # Optional: keep or drop str_suppl3 and po_box
            # df.drop(columns=['address_house_no', 'address_house_no2', 'po_box', 'address_line2'], inplace=True, errors='ignore')

        elif manufacturer in ["bmw", "mini"]:
            if 'address_1' in df.columns or 'address_2' in df.columns:
                part1 = df.get('address_1', '').fillna('')
                part2 = df.get('address_2', '').fillna('')
                df['address_line1'] = (part1 + ' ' + part2).str.replace(r'\s+', ' ', regex=True).str.strip()
                self.log("BMW/MINI: Combined address_1 and address_2 into address_line1")

        elif manufacturer in ["nissan", "infiniti", "mercedes"]:
            if 'address_line1' in df.columns and 'address_line2' in df.columns:
                df['address_line1'] = (
                    df['address_line1'].fillna('') + ' ' + df['address_line2'].fillna('')
                ).str.replace(r'\s+', ' ', regex=True).str.strip()
                self.log(f"{manufacturer.upper()}: Combined address_line1 + address_line2")

        elif manufacturer in ["honda", "acura"]:
            # Already mapped directly as address_line1 — no issues
            self.log(f"{manufacturer.upper()}: Address already in address_line1")

        elif manufacturer == "volvo":
            # Volvo columns are positional but already mapped to address_line1
            self.log("Volvo: Positional mapping handled via column map")

        return df




    def step_deduplicate_addresses(self):
        try:
            input_path = Path(self.output_entry.get()) / "email_cleaned.csv"
            output_path = Path(self.output_entry.get()) / "after_address_dedup.csv"
            filtered_out_path = Path(self.output_entry.get()) / "filtered_out_duplicate_address.csv"

            # Load data
            df = pd.read_csv(input_path, dtype=str)

            # --- Standardize fields for robust matching (upper, strip, fillna) ---
            df['naddrs1'] = df['address_line1'].fillna('').str.upper().str.strip()
            df['ncity']   = df['city'].fillna('').str.upper().str.strip()
            df['nstate']  = df['state'].fillna('').str.upper().str.strip()

            # Sort for reproducibility (optional: sort by all fields)
            df = df.sort_values(['naddrs1', 'ncity', 'nstate'])

            # Find duplicates: keep the first occurrence (address+city+state combo)
            dup_mask = df.duplicated(subset=['naddrs1', 'ncity', 'nstate'], keep='first')

            # Filtered duplicates
            duplicates = df[dup_mask].copy()
            duplicates['drop_reason'] = 'Filter_Duplicate_Address'

            # Non-duplicates
            df_unique = df[~dup_mask].copy()

            # Save filtered out duplicates for audit
            duplicates.to_csv(filtered_out_path, index=False)

            # Save deduplicated data for next step
            df_unique.to_csv(output_path, index=False)

            self.log(f"Duplicate address filter: removed {len(duplicates)} duplicates; {len(df_unique)} records remain.")
            self.tracker.set('duplicate_addresses', len(duplicates))

            return True

        except Exception as e:
            self.log(f"Error in duplicate address filter: {str(e)}")
            return False




    def get_zip_column(self, df):
        # Normalize column names to lowercase for safe matching
        lowercase_cols = [col.lower() for col in df.columns]
        zip_candidates = ['zip_code', 'zipcode', 'zip']

        for candidate in zip_candidates:
            if candidate in lowercase_cols:
                # Get original column name by position/index
                matched_col = df.columns[lowercase_cols.index(candidate)]
                self.log(f"Matched ZIP column: {matched_col}")
                return matched_col

        self.log(f"No valid ZIP column found. Available columns: {df.columns.tolist()}")
        return None



    def normalize_zip_column(self, df, zip_col):
        if zip_col not in df.columns:
            self.log(f"normalize_zip_column: Column '{zip_col}' not found.")
            return df

        # Step 1: Clean ZIP content
        df[zip_col] = df[zip_col].astype(str).str.strip()
        df[zip_col] = df[zip_col].replace(['nan', 'NaN', np.nan], '')

        # Step 2: Remove rows with non-numeric ZIPs
        non_numeric = ~df[zip_col].str.match(r'^\d+$')
        if non_numeric.any():
            self.log(f"normalize_zip_column: Removed {non_numeric.sum()} non-numeric ZIPs from '{zip_col}'")
            df.loc[non_numeric, zip_col] = ''

        # Step 3: Pad ZIPs to length 5
        df[zip_col] = df[zip_col].str.zfill(5).str[:5]

        # Step 4: Create zip2 (2-digit prefix) based on validated ZIP
        df['zip2'] = df[zip_col].astype(str).str[:2]
        self.log(f"ZIP normalization successful using column '{zip_col}'")

        # Step 5: Log sample ZIPs
        zip_sample = df[zip_col].dropna().head(5).tolist()
        self.log(f"normalize_zip_column: Sample standardized ZIPs in '{zip_col}': {zip_sample}")

        return df


        
        """
        Assigns purch/lse based on sales type rules:
        - For Mercedes: 2 for lease, 1 for purchase/retail, reject if missing/invalid.
        - For Volvo and others: if no sales type column, set purch/lse = 1 for all.
        """
    def step_sales_type_filter(self):
        try:
            input_path = Path(self.output_entry.get()) / "after_address_dedup.csv"
            df = pd.read_csv(input_path, dtype=str)
            df = self.standardize_columns(df)
            manufacturer = self.manufacturer_combo.get().lower().strip()

            # Case-insensitive match for known sales type columns
            sales_type_col = None
            lowered_columns = {col.lower(): col for col in df.columns}
            for candidate in ['veh_acq_type_cd', 'purchase_or_lease', 'sales_type']:
                if candidate in lowered_columns:
                    sales_type_col = lowered_columns[candidate]
                    break

            if manufacturer == "mercedes" and sales_type_col:
                df[sales_type_col] = df[sales_type_col].str.strip().str.lower()
                allowed_types = {'lease', 'retail', 'purchase'}
                mask_valid = df[sales_type_col].isin(allowed_types)
                rejected = df[~mask_valid].copy()
                accepted = df[mask_valid].copy()
                accepted['purch/lse'] = accepted[sales_type_col].apply(lambda x: 2 if x == 'lease' else 1)
                accepted['purch/lse'] = accepted['purch/lse'].astype(int)
                self.tracker.set('invalid_sales_type', len(rejected))
                if not rejected.empty:
                    rejected['rejection_reason'] = 'Invalid or missing sales type'
                    rejected.to_csv(Path(self.output_entry.get()) / "sales_type_rejected.csv", index=False)
                self.df = accepted
                self.log(f"Sales Type Filter: {len(rejected)} records rejected, {len(self.df)} accepted")
            else:
                df['purch/lse'] = 1
                df['purch/lse'] = df['purch/lse'].astype(int)
                self.df = df
                self.log(f"Sales Type Filter: No usable sales type column found or manufacturer != Mercedes. Set purch/lse = 1 for all {len(df)} records.")

            # --- WRITE filtered dataframe for subsequent use ---
            output_path = Path(self.output_entry.get()) / "sales_filtered.csv"
            self.df.to_csv(output_path, index=False)
            self.log(f"Sales Type Filter: Wrote {len(self.df)} accepted records to {output_path}")

            return True

        except Exception as e:
            self.log(f"Sales Type Filter failed: {str(e)}")
            return False



    def step_geography_filter(self):
        try:
            input_path = Path(self.output_entry.get()) / "sales_filtered.csv"

            if not input_path.exists() or os.path.getsize(input_path) == 0:
                self.log(f"Geography Filter: Input file missing or empty: {input_path}")
                return False

            output_path = Path(self.output_entry.get()) / "geo_filtered.csv"

            df = pd.read_csv(input_path, dtype=str)
            self.log(f"Geography Filter: Starting with {len(df)} records")

            # --- Ensure 'state' column exists ---
            if 'state' not in df.columns:
                df['state'] = ''
                self.log(f"WARNING: Added missing column: state")
            df['state'] = df['state'].str.strip().str.upper()

            # --- ZIP Standardization ---
            zip_col = self.get_zip_column(df)
            if zip_col:
                df = self.normalize_zip_column(df, zip_col)
                self.log(f"✅ ZIP normalization using column '{zip_col}' completed.")
            else:
                self.log("⚠ No ZIP column found — skipping ZIP normalization and geography filtering step.")
                return False  # ⛔ Prevents KeyError on zip2 access later

            # --- Track Metrics ---
            bad_zip_count = (df[zip_col] == '').sum() if zip_col else 0
            self.tracker.set('blank_or_invalid_zip', bad_zip_count)

            # Debugging Samples
            if zip_col:
                sample_padded = df[zip_col].dropna().sample(min(5, len(df))).tolist()
                self.log(f"Sample ZIP conversions in {zip_col}: {sample_padded}")

            # --- Load Validation Rules (REQUIRED) ---
            rules_path = self.config_dir / "state_zip_rules.yml"
            if not rules_path.exists():
                raise FileNotFoundError(f"State/ZIP rules file missing: {rules_path}")

            with open(rules_path, encoding="utf-8") as f:
                rules = yaml.safe_load(f)
            valid_combinations = rules['valid_combinations']

            # --- Geography Validation Mask ---
            mask = pd.Series(False, index=df.index)
            for state, z2list in valid_combinations.items():
                state_mask = (df['state'] == state)
                zip_mask = df['zip2'].isin(z2list)
                mask |= (state_mask & zip_mask)

            # --- Filter Records ---
            valid_geo = df[mask]
            rejected_geo = df[~mask]

            # --- Breakdown Rejections ---
            bad_state_mask = ~rejected_geo['state'].isin(valid_combinations.keys())
            bad_state_codes = rejected_geo[bad_state_mask].copy()
            bad_state_codes['rejection_reason'] = 'Invalid state code'

            bad_zip_state = rejected_geo[~bad_state_mask].copy()
            bad_zip_state['rejection_reason'] = 'State/ZIP mismatch'

            self.log(f"Geography Filter: Rejected {len(bad_state_codes)} bad states, {len(bad_zip_state)} state/ZIP mismatches")
            self.log(f"Geography Filter: {len(valid_geo)} valid records remain")

            # --- Metrics ---
            self.tracker.set('bad_state_codes', len(bad_state_codes))
            self.tracker.set('bad_zip_state', len(bad_zip_state))

            # --- Save Split Outputs ---
            Path(self.output_entry.get()).mkdir(parents=True, exist_ok=True)

            # Save rejected records
            pd.concat([bad_state_codes, bad_zip_state]).to_csv(
                Path(self.output_entry.get()) / "geo_rejected.csv",
                index=False,
                encoding='utf-8'
            )

            # Sanitize and save valid records
            valid_geo = valid_geo.applymap(
                lambda x: x.encode('utf-8', 'ignore').decode('utf-8') if isinstance(x, str) else x
            )
            valid_geo.to_csv(output_path, index=False, encoding='utf-8')  # ✅ This line was missing

            return True

        except Exception as e:
            self.log(f"Geography filter failed: {str(e)}")
            return False



    def step_business_filter(self):
        import pandas as pd
        import chardet
        from pathlib import Path

        try:
            input_path = Path(self.output_entry.get()) / "geo_filtered.csv"
            output_path = Path(self.output_entry.get()) / "business_filtered.csv"
            
            # --- Try all encodings and record errors for debugging ---
            encodings_to_try = ['utf-8', 'latin-1', 'cp1252']
            df = None
            first_bytes = None
            error_at_line = None
            last_exception = None

            for enc in encodings_to_try:
                try:
                    self.log(f"Trying to read {input_path} with encoding: {enc}")
                    df = pd.read_csv(input_path, dtype=str, encoding=enc)
                    self.log(f"SUCCESS: {input_path} loaded fine with {enc}.")
                    break
                except Exception as e:
                    self.log(f"FAIL: {input_path} with {enc}: {e}")
                    last_exception = e

            # If still no dataframe, do raw scan and chardet guess
            if df is None:
                with open(input_path, "rb") as f:
                    first_bytes = f.read(4096)
                    guess = chardet.detect(first_bytes)
                self.log(f"chardet guess for file: {guess}")
                
                # Try row-by-row binary scan to find problematic lines
                bad_lines = []
                with open(input_path, "rb") as f:
                    for i, line in enumerate(f, 1):
                        try:
                            line.decode('utf-8')
                        except Exception as ue:
                            self.log(f"Row {i}: UTF-8 decode failed: {ue}. Bytes: {repr(line[:40])}...")
                            bad_lines.append((i, line[:80]))
                            error_at_line = i
                            break  # Stop at first problem for now
                if bad_lines:
                    self.log(f"First bad line: Row {bad_lines[0][0]}, Bytes {bad_lines[0][1]!r}")
                try:
                    df = pd.read_csv(input_path, dtype=str, encoding=guess.get('encoding') or 'cp1252')
                    self.log(f"LOADED with chardet/guess encoding: {guess.get('encoding') or 'cp1252'}")
                except Exception as e2:
                    self.log(f"ALL encoding attempts failed. Last: {e2}")
                    self.log(f"First 80 file bytes: {first_bytes[:80]!r}")
                    msg = (
                        f"Cannot read {input_path}.\n"
                        f"Last parsing error: {e2}\n"
                        f"Likely non-UTF-8/non-Latin1 character on line {error_at_line}.\n"
                        f"Sample bad bytes: {bad_lines[0][1]!r}" if bad_lines else ""
                    )
                    messagebox.showerror("File Encoding Error", msg, parent=self.root)
                    return False

            # === Business logic below unchanged ===
            df = self.standardize_columns(df)
            before = len(df)
            self.log(f"Business Filter: Starting with {before} records")

            business_terms_path = self.config_dir / "business_terms.yml"
            with open(business_terms_path, encoding="utf-8") as f:
                business_terms = yaml.safe_load(f)

            RISKY_TERMS = {"CLUB", "UCC", "DIVISION", "COLLEGE"}
            all_terms = business_terms['business_indicators']
            riskier_terms = [t for t in all_terms if t.strip().upper() in RISKY_TERMS]
            safe_terms    = [t for t in all_terms if t.strip().upper() not in RISKY_TERMS]

            risky_patterns = [re.compile(re.escape(term), re.IGNORECASE) for term in riskier_terms]
            safe_patterns  = [re.compile(term, re.IGNORECASE) for term in safe_terms]
            business_indicator_mask = pd.Series(False, index=df.index)

            # In NAME
            if 'name' in df.columns:
                name_hits = df['name'].apply(
                    lambda x: self.matches_any_pattern(x, risky_patterns + safe_patterns) if pd.notnull(x) else False
                )
                business_indicator_mask |= name_hits
                self.log(f"Business filter: {name_hits.sum()} rows hit business indicator in 'name'.")

            oem_mask = pd.Series(False, index=df.index)
            if {'first_name', 'last_name'}.issubset(df.columns):
                full_name = (df['first_name'].str.lower().fillna('') + ' ' + df['last_name'].str.lower().fillna(''))
                for oem in business_terms['oem_manufacturers']:
                    current_oem_match = full_name.str.contains(rf'\b{oem.lower()}\b', regex=True, na=False)
                    hits = current_oem_match.sum()
                    self.log(f"OEM business filter: '{oem}' found in {hits} rows")
                    oem_mask |= current_oem_match

            business_mask = oem_mask | business_indicator_mask
            self.log(f"Flagged as possible business: {business_mask.sum()} / {before}")

            # Exclusions
            for exclusion in business_terms.get('hardcoded_exclusions', []):
                name_match = df['name'].str.strip().str.lower().fillna('') == exclusion['name'].strip().lower()
                if 'city' in exclusion and 'city' in df.columns:
                    city_match = df['city'].str.strip().str.lower().fillna('') == exclusion['city'].strip().lower()
                    exclusion_mask = name_match & city_match
                else:
                    exclusion_mask = name_match
                business_mask = business_mask & ~exclusion_mask

            self.log(f"After exclusions: {business_mask.sum()} business records / {before}")

            valid_records = df[~business_mask]
            rejected_business = df[business_mask]
            after = len(valid_records)
            self.tracker.log_step('Business Filter', before, len(rejected_business), after)
            self.tracker.set('business_exclusions', len(rejected_business))

            rejected_business.to_csv(Path(self.output_entry.get()) / "business_rejected.csv", index=False, encoding="utf-8")
            valid_records.to_csv(output_path, index=False, encoding="utf-8")
            self.log(f"Business Filter: Rejected {len(rejected_business)}. Remaining: {after}.")

            return True

        except Exception as e:
            import traceback
            self.log(f"Business filter failed: {str(e)}\n{traceback.format_exc()}")
            return False






    # Ensure this helper method is in the class:
    def matches_any_pattern(self, text, patterns):
        """Check if text matches any regex pattern with NaN safety"""
        if pd.isna(text) or text is None:
            return False
        return any(p.search(str(text)) for p in patterns)




    def step_vin_processing(self):
        import traceback
        import pandas as pd
        from pathlib import Path
        import yaml
        import logging
        
        try:
            # Debug: Start of step
            self.log("==== VIN PROCESSING STEP STARTED ====")
            manufacturer = self.manufacturer_combo.get().lower().strip()  # ✅ define it here

            input_path = Path(self.output_entry.get()) / "business_filtered.csv"
            output_path = Path(self.output_entry.get()) / "vin_processed.csv"
            output_dir = output_path.parent
            output_dir.mkdir(parents=True, exist_ok=True)

            # 1. Read file with debug
            self.log(f"Reading file: {input_path}")
            df = pd.read_csv(input_path, dtype={'vin': str})
            self.log(f"Initial shape: {df.shape}, Columns: {list(df.columns)}")
            self.log(f"VIN dtype: {df['vin'].dtype}")
            
            # 2. Standardize columns
            df = self.standardize_columns(df)
            self.log(f"After standardization - Columns: {list(df.columns)}")

            # Fallback: ensure bodycode exists for brands like JLR
            if 'bodycode' not in df.columns and 'modelid' in df.columns:
                df['bodycode'] = df['modelid']
                self.log("Fallback: Populated 'bodycode' from 'modelid' (JLR compatibility)")

            # 3. Pre-cleaning debug
            self.log("\n==== PRE-CLEANING DEBUG ====")
            self.log(f"VIN types: {df['vin'].apply(type).value_counts().to_dict()}")
            self.log(f"Sample VINs (first 10): {df['vin'].head(10).tolist()}")
            
            # Check for problematic values
            non_string_vins = df[~df['vin'].apply(lambda x: isinstance(x, str))]
            if not non_string_vins.empty:
                self.log(f"WARNING: Found {len(non_string_vins)} non-string VINs")
                self.log(f"Sample non-string VINs: {non_string_vins['vin'].head().tolist()}")
            
            # 4. VIN cleaning
            self.log("\n==== CLEANING VIN COLUMN ====")
            df['vin'] = df['vin'].astype(str)
            df['vin'] = (
                df['vin']
                .str.replace(r'\.0$', '', regex=True)
                .str.strip()
                .str.upper()
                .str.replace(r'[^A-Z0-9]', '', regex=True)
            )
            
            # 5. Post-cleaning debug
            self.log("\n==== POST-CLEANING DEBUG ====")
            self.log(f"VIN types: {df['vin'].apply(type).value_counts().to_dict()}")
            self.log(f"Sample VINs (first 10): {df['vin'].head(10).tolist()}")
            self.log(f"VIN lengths: {df['vin'].str.len().value_counts().to_dict()}")
            
            # 6. VIN validation
            self.log("\n==== VIN VALIDATION ====")
            mask_valid_vin = (
                df['vin'].ne('') & 
                df['vin'].apply(lambda x: len(x) == 17) & 
                df['vin'].apply(str.isalnum)
            )
            self.log(f"Validation mask: {mask_valid_vin.sum()} valid, {len(df)-mask_valid_vin.sum()} invalid")
            
            # Handle validation
            valid_vins = df[mask_valid_vin].copy()
            rejected_vins = df[~mask_valid_vin].copy()
            rejected_vins['rejection_reason'] = 'Invalid VIN'
            rejected_vins.to_csv(output_dir / "vin_rejected.csv", index=False)
            self.tracker.log_step('VIN Validation', len(df), len(rejected_vins), len(valid_vins))
            self.tracker.set('vin_rejected', len(rejected_vins))
            df = valid_vins

            # VIN Mapping
            vin_map_path = self._get_vin_mapping_path(manufacturer, self.vin_mapping_dir)
            with open(vin_map_path, encoding="utf-8") as f:
                vin_mapping = yaml.safe_load(f)

            # Define default cell code NOW (so it's available to functions below)
            default_cell = vin_mapping.get('default_cell_code', 'DEFAULT')
            
            # Extract VIN segments
            df['vin13'] = df['vin'].str[:3].str.upper()
            df['vin48'] = df['vin'].str[3:8].str.upper()
            df['vin10'] = df['vin'].str[9].str.upper()

            # ✅ Add this mapping after vin13 is set
            WMI_TO_BRAND = {
                "SAL": "Land Rover",
                "SAJ": "Jaguar"
            }
            df['true_brand'] = df['vin13'].map(WMI_TO_BRAND).fillna(df.get('brand', ''))

            self.log("Example mapped brands: " + str(df[['vin13', 'true_brand']].dropna().head(5).to_dict('records')))


            # Model year mapping
            VIN10_TO_YEAR = {
                'A': 2010, 'B': 2011, 'C': 2012, 'D': 2013, 'E': 2014, 'F': 2015, 
                'G': 2016, 'H': 2017, 'J': 2018, 'K': 2019, 'L': 2020, 'M': 2021,
                'N': 2022, 'P': 2023, 'R': 2024, 'S': 2025, 'T': 2026, 'V': 2027,
                'W': 2028, 'X': 2029, 'Y': 2030, '1': 2001, '2': 2002, '3': 2003,
                '4': 2004, '5': 2005, '6': 2006, '7': 2007, '8': 2008, '9': 2009
            }
            df['model_year'] = df['vin10'].map(VIN10_TO_YEAR)
            valid_years = [2023, 2024, 2025, 2026]
            valid_mask = df['model_year'].isin(valid_years)
            invalid_years = df[~valid_mask]
            if not invalid_years.empty:
                invalid_years['rejection_reason'] = 'Invalid model year'
                invalid_years.to_csv(
                    Path(self.output_entry.get()) / "vin_rejected.csv",
                    mode='a', header=False, index=False
                )
                self.tracker.set('invalid_model_year', len(invalid_years))
            df = df[valid_mask]

            # Add modelid if missing
            if 'modelid' not in df.columns:
                df['modelid'] = ''

            # Define cell code assignment with conflict resolution
            def get_cell_code(row, rules):
                vin13 = row['vin13']
                vin48 = row['vin48']
                vin10 = row['vin10']

                # CORRECTED: Use bodycode instead of modelid
                source_bodycode = row.get('bodycode', '')
                bodycode_str = str(source_bodycode).strip() if pd.notna(source_bodycode) else ''
                
                # 1. First check: Match VIN segments AND bodycode (if both exist)
                if bodycode_str:
                    matching_rules = [r for r in rules 
                                    if r.get('vin13') == vin13 
                                    and r.get('vin48') == vin48 
                                    and r.get('vin10') == vin10 
                                    and r.get('modelid', '').strip() == bodycode_str]
                    if matching_rules:
                        return matching_rules[0]['cell']
                
                # 2. Second check: Match VIN segments only (FIXED LOGIC)
                matching_rules = [r for r in rules 
                                if r.get('vin13') == vin13 
                                and r.get('vin48') == vin48 
                                and r.get('vin10') == vin10 
                                and (not r.get('modelid') or r.get('modelid', '').strip() == '')]
                if matching_rules:
                    return matching_rules[0]['cell']
                
                # 3. Fallback to default if no match
                return vin_mapping.get('default_cell_code', 'DEFAULT')
 
 
            def lookup_vin_details(row, rules):
                vin13 = str(row['vin13']).strip().upper()
                vin48 = str(row['vin48']).strip().upper()
                vin10 = str(row['vin10']).strip().upper()
                bodycode = str(row.get('bodycode', '')).strip().upper()

                # 1. Match VIN segments + modelid/bodycode
                for r in rules:
                    if (str(r.get('vin13', '')).strip().upper() == vin13 and
                        str(r.get('vin48', '')).strip().upper() == vin48 and
                        str(r.get('vin10', '')).strip().upper() == vin10 and
                        str(r.get('modelid', '')).strip().upper() == bodycode):
                        return r.get('cell'), r.get('brand'), r.get('desc'), r.get('modelid')

                # 2. Match VIN segments only (modelid empty)
                for r in rules:
                    if (str(r.get('vin13', '')).strip().upper() == vin13 and
                        str(r.get('vin48', '')).strip().upper() == vin48 and
                        str(r.get('vin10', '')).strip().upper() == vin10 and
                        (not r.get('modelid') or str(r.get('modelid', '')).strip() == '')):
                        return r.get('cell'), r.get('brand'), r.get('desc'), r.get('modelid')

                # 3. Fallback
                return default_cell, '', '', ''   # Use local default_cell, not vin_mapping.get...

            df[['cell', 'brand', 'desc', 'modelid']] = df.apply(
                lambda row: pd.Series(lookup_vin_details(row, vin_mapping['vin_mappings'])),
                axis=1
            )

            # Missing cell code report/logging (moved after default_cell assignment)
            unmatched = df[df['cell'] == default_cell]
            if not unmatched.empty:
                self.log(f"WARNING: {len(unmatched)} records returned default cell codes.")
                self.log("Sample unmatched VINs:\n" +
                    str(unmatched[['vin13', 'vin48', 'vin10', 'bodycode']].head().to_dict(orient='records')))


            # Debug logging
            self.log(f"First 5 cell codes: {df['cell'].head().tolist()}")

            # ==== Missing Cell Handling and Reporting ====
            # Define default cell code
            default_cell = vin_mapping.get('default_cell_code', 'DEFAULT')

            # Ensure cell column exists and fill missing values
            if 'cell' in df.columns:
                df['cell'].fillna(default_cell, inplace=True)
                # Check for any remaining missing values
                missing_count = df['cell'].isna().sum()
                if missing_count > 0:
                    self.log(f"WARNING: {missing_count} records still have missing cell codes")
            else:
                df['cell'] = default_cell
                self.log("Created 'cell' column with default values")

            # Create missing cell report WITH BODYCODE
            missing_cell_df = df[df['cell'] == default_cell]
            missing_cell_count = len(missing_cell_df)
            if missing_cell_count > 0:
                report_columns = [
                    'vin', 'vin13', 'vin48', 'vin10', 'bodycode', 'cell',
                    'model', 'trimline',  # ✅ Add these here
                    'first_name', 'last_name', 'address_line1', 'city', 'state', 'zip_code',
                    'brand', 'desc', 'model_year'
                ]
                # Only include existing columns
                available_columns = [col for col in report_columns if col in missing_cell_df.columns]
                missing_cell_path = output_dir / "missing_cell_codes.csv"
                missing_cell_df.to_csv(missing_cell_path, columns=available_columns, index=False)
                self.log(f"Saved {missing_cell_count} missing cell records to {missing_cell_path}")
 
                sample_cols = [col for col in ['model', 'trimline'] if col in missing_cell_df.columns]
                if sample_cols:
                    self.log(f"Sample missing cell data: {missing_cell_df[sample_cols].dropna().head().to_dict('records')}")
                else:
                    self.log("No 'model' or 'trimline' columns found in missing_cell_df.")

            # Ensure critical columns exist
            for col in ['brand', 'desc']:
                if col not in df.columns:
                    df[col] = ''
                    self.log(f"WARNING: Added missing {col} column")

            # Final validation
            if 'cell' not in df.columns:
                raise ValueError("'cell' column missing after VIN mapping")

            # Save final output
            df.to_csv(output_path, index=False)
            self.log(f"VIN Processing: Saved {len(df)} valid records")
            return True
           
        except Exception as e:
            self.log(f"CRITICAL ERROR in VIN processing: {str(e)}")
            self.log(traceback.format_exc())
            return False


    def step_deduplication(self):
        try:
            input_path = Path(self.output_entry.get()) / "vin_processed.csv"
            output_path = Path(self.output_entry.get()) / "deduped.csv"
            df = pd.read_csv(input_path, dtype=str)
            df = self.standardize_columns(df)

            # --- Metrics Tracking ---
            before = len(df)

            # --- Column Check ---
            if 'cell' not in df.columns or 'model_year' not in df.columns:
                raise ValueError(f"Deduplication input missing columns. Columns: {df.columns.tolist()}")

            # --- Metrics: Before deduplication ---
            initial_count = len(df)
            self.log(f"Deduplication: Starting with {initial_count} records")
            self.tracker.set('dedup_initial', initial_count)

            # --- Load ALL VINs (historical + TrueCar) ---
            all_vins = set()
            
            # Process HISTORICAL files from listbox
            historical_files = self.hist_listbox.get(0, tk.END)
            for hist_path in historical_files:
                hist_path = Path(hist_path.strip())
                if not hist_path.exists():
                    self.log(f"Skipping missing historical file: {hist_path}")
                    continue
                    
                try:
                    if hist_path.suffix.lower() in ('.xls', '.xlsx'):
                        hist_df = pd.read_excel(hist_path, dtype=str, engine='openpyxl')
                    else:
                        hist_df = pd.read_csv(hist_path, dtype=str)
                    
                    hist_df = self.standardize_columns(hist_df)
                    vin_col = next((col for col in hist_df.columns 
                                            if col.strip().lower() in self.VIN_COLUMN_NAMES), None)
                    if not vin_col:
                        self.log(f"No VIN column found in {hist_path.name}. Columns present: {list(hist_df.columns)}")

                    if vin_col:
                        vins = hist_df[vin_col].dropna().astype(str).str.upper()
                        all_vins.update(vins)
                        self.log(f"Loaded {len(vins)} VINs from {hist_path.name}")
                    else:
                        self.log(f"No VIN column found in {hist_path.name}")
                except Exception as e:
                    self.log(f"Error processing {hist_path.name}: {str(e)}")

            # Process TRUECAR files from listbox
            truecar_files = self.truecar_listbox.get(0, tk.END)
            for tc_path in truecar_files:
                tc_path = Path(tc_path.strip())
                if not tc_path.exists():
                    self.log(f"Skipping missing TrueCar file: {tc_path}")
                    continue
                    
                try:
                    if tc_path.suffix.lower() in ('.xls', '.xlsx'):
                        tc_df = pd.read_excel(tc_path, dtype=str, engine='openpyxl')
                    else:
                        tc_df = pd.read_csv(tc_path, dtype=str, sep=None, engine='python')
                        
                    tc_df = self.standardize_columns(tc_df)
                    vin_col = next((col for col in tc_df.columns 
                                    if col.strip().lower() in self.VIN_COLUMN_NAMES), None)
                    if not vin_col:
                        self.log(f"No VIN column found in {tc_path.name}. Columns present: {list(tc_df.columns)}")

                    if vin_col:
                        vins = tc_df[vin_col].dropna().astype(str).str.upper()
                        all_vins.update(vins)
                        self.log(f"Loaded {len(vins)} VINs from {tc_path.name}")
                    else:
                        self.log(f"No VIN column found in {tc_path.name}")
                except Exception as e:
                    self.log(f"Error processing TrueCar file {tc_path.name}: {str(e)}")

            # --- Create Duplicate Mask ---
            duplicate_mask = df['vin'].str.upper().isin(all_vins) if 'vin' in df.columns else pd.Series([False] * len(df))

            # Add internal deduplication
            self.log(f"Pre-internal-dedup: {len(df)} records")
            df = df.drop_duplicates(subset=['vin'], keep='first')  # Critical fix
            self.log(f"Post-internal-dedup: {len(df)} records")

            # After internal deduplication
            internal_dupes = before - len(df)
            self.log(f"Internal duplicates removed: {internal_dupes}")


            # --- Split Records ---
            deduped = df[~duplicate_mask]
            duplicates = df[duplicate_mask]

            after = len(deduped)
            self.tracker.log_step('Deduplication', before, len(duplicates), after)

            # --- Save Rejected Duplicates ---
            duplicates = duplicates.copy()
            duplicates['rejection_reason'] = 'Duplicate record'
            duplicates.to_csv(
                Path(self.output_entry.get()) / "duplicates_rejected.csv",
                index=False
            )

            # --- Save Deduped Data ---
            deduped.to_csv(output_path, index=False)

            # --- Log & Track ---
            self.log(f"Deduplication: {len(duplicates)} records removed as duplicates, {len(deduped)} remain")
            self.tracker.set('dedup_removed', len(duplicates))
            self.tracker.set('dedup_remaining', len(deduped))
            self.status(f"Deduplication completed: {len(deduped)} valid, {len(duplicates)} duplicates removed.")

            return True
        except Exception as e:
            self.log(f"Deduplication failed: {str(e)}")
            return False



    def step_ucc_check(self):
        try:
            input_path = Path(self.output_entry.get()) / "deduped.csv"
            ucc_path = self.ucc_entry.get()

            # Read current data
            current_df = pd.read_csv(input_path, dtype=str)
            current_df = self.standardize_columns(current_df)
            current_df.columns = [col.strip().lower() for col in current_df.columns]

            # Rename model_year/year to modyy
            rename_occurred = False
            if 'model_year' in current_df.columns:
                self.log("Renamed 'model_year' column to 'modyy' for UCC check")
                current_df.rename(columns={'model_year': 'modyy'}, inplace=True)
                rename_occurred = True
            elif 'year' in current_df.columns:
                self.log("Renamed 'year' column to 'modyy' for UCC check")
                current_df.rename(columns={'year': 'modyy'}, inplace=True)
                rename_occurred = True
            
            # Remove duplicate columns after renaming
            current_df = current_df.loc[:, ~current_df.columns.duplicated()]

            # --- Read UCC file robustly ---
            if ucc_path.lower().endswith(('.xls', '.xlsx')):
                ucc_df = pd.read_excel(ucc_path, sheet_name=0, dtype=str)
                ucc_df = ucc_df.dropna(how='all').dropna(axis=1, how='all')
            else:
                try:
                    ucc_df = pd.read_csv(ucc_path, dtype=str, encoding='utf-8', on_bad_lines='skip')
                except UnicodeDecodeError:
                    ucc_df = pd.read_csv(ucc_path, dtype=str, encoding='latin-1', on_bad_lines='skip')
            
            ucc_df = self.standardize_columns(ucc_df)
            ucc_df.columns = [col.strip().lower() for col in ucc_df.columns]

            if ucc_df.empty:
                raise ValueError("UCC file is empty after cleaning")

            # --- Normalize and Rename Columns ---
            rename_map = {
                'sra cell': 'cell',
                'year': 'modyy',
                'model year': 'modyy',
                'modelyear': 'modyy',
                '21 digit': 'ucc',
            }
            ucc_df.rename(columns=rename_map, inplace=True)
            
            # Remove duplicate columns after renaming
            ucc_df = ucc_df.loc[:, ~ucc_df.columns.duplicated()]
            self.log(f"UCC columns: {ucc_df.columns.tolist()}")
            
            if 'ucc' not in ucc_df.columns:
                raise ValueError(f"UCC column missing. Columns: {ucc_df.columns.tolist()}")

            # --- Validate required columns ---
            required_cols = {'cell', 'modyy'}
            for df, name in [(current_df, "Current Data"), (ucc_df, "UCC File")]:
                missing = required_cols - set(df.columns)
                if missing:
                    self.log(f"{name} missing columns: {missing}")
                    self.log(f"{name} columns: {list(df.columns)}")
                    raise ValueError(f"{name} missing required columns: {missing}")

            # --- Clean Data Types for Matching ---
            def clean_series(data):
                """Handles both Series and DataFrame inputs safely"""
                if isinstance(data, pd.DataFrame):
                    data = data.iloc[:, 0]
                return data.astype(str).str.strip().str.upper()
            
            # Clean UCC columns
            ucc_columns = ['cell', 'modyy', 'ucc', 'description']
            for df in [current_df, ucc_df]:
                for col in ucc_columns:
                    if col in df.columns:
                        # Convert to Series if needed
                        if isinstance(df[col], pd.DataFrame):
                            df[col] = df[col].squeeze()
                        # Clean the series
                        df[col] = clean_series(df[col])

            # Log shapes and samples
            self.log(f"Current data shape: {current_df.shape}")
            self.log(f"UCC data shape: {ucc_df.shape}")
            self.log(f"Sample current_df cell/modyy: {current_df[['cell', 'modyy']].head().to_dict('records')}")
            self.log(f"Sample ucc_df cell/modyy: {ucc_df[['cell', 'modyy']].head().to_dict('records')}")

            # Add type checks
            cell_type = type(current_df['cell'])
            modyy_type = type(current_df['modyy'])
            self.log(f"Type check - cell: {cell_type}, modyy: {modyy_type}")

            # --- Debugging ---
            # Type and invisible character checks
            self.log(f"Data types - current_df: cell={current_df['cell'].dtype}, modyy={current_df['modyy'].dtype}")
            self.log(f"Data types - ucc_df: cell={ucc_df['cell'].dtype}, modyy={ucc_df['modyy'].dtype}")
            
            def log_invisible_chars(series, name):
                samples = series.head(3).apply(lambda x: repr(x))
                self.log(f"Invisible chars in {name}: {samples.tolist()}")
            
            log_invisible_chars(current_df['cell'], "current cell")
            log_invisible_chars(ucc_df['cell'], "UCC cell")
            
            # --- Combination Check ---
            # Ensure no duplicate columns exist
            if 'cell' in current_df.columns and 'modyy' in current_df.columns:
                current_combos = current_df[['cell', 'modyy']].drop_duplicates()
            else:
                raise ValueError("'cell' or 'modyy' columns missing in current_df")
                
            if 'cell' in ucc_df.columns and 'modyy' in ucc_df.columns:
                ucc_combos = ucc_df[['cell', 'modyy']].drop_duplicates()
            else:
                raise ValueError("'cell' or 'modyy' columns missing in ucc_df")
            
            merged = pd.merge(
                current_combos, 
                ucc_combos, 
                on=['cell', 'modyy'], 
                how='left', 
                indicator=True
            )
            missing_combos = merged[merged['_merge'] == 'left_only'][['cell', 'modyy']]
            
            if not missing_combos.empty:
                missing_path = Path(self.output_entry.get()) / "missing_ucc_combinations.csv"
                missing_combos.to_csv(missing_path, index=False)
                self.log(f"WARNING: Found {len(missing_combos)} missing UCC combinations")
                self.log(f"Saved missing combos to {missing_path}")
            else:
                self.log("All cell/year combinations present in UCC file")

            # Safe value extraction
            def get_first_value(column):
                if isinstance(column, pd.Series):
                    return column.iloc[0]
                elif isinstance(column, pd.DataFrame):
                    return column.iloc[0, 0]
                return column.iloc[0]

            sample_cell = str(get_first_value(current_df['cell']))
            sample_modyy = str(get_first_value(current_df['modyy']))

            # Robust comparison
            match = ucc_df[
                ucc_df['cell'].astype(str).eq(sample_cell) & 
                ucc_df['modyy'].astype(str).eq(sample_modyy)
            ]

            self.log(f"Sample match: Cell={sample_cell}, Year={sample_modyy} -> {'FOUND' if not match.empty else 'MISSING'}")

            # --- Always save and continue ---
            output_path = Path(self.output_entry.get()) / "ucc_checked.csv"
            current_df.to_csv(output_path, index=False)
            self.log(f"Saved UCC-checked data to {output_path}")
            return True  # Always continue

        except Exception as e:
            self.log(f"UCC check failed: {str(e)}")
            self.log(traceback.format_exc())
            self.status("UCC check failed.")
            return False




    def step_ucc_update(self):
        try:
            # Validate missing entries
            if not hasattr(self, 'missing_entries') or self.missing_entries.empty:
                self.log("No missing UCC entries to update")
                self.status("No UCC updates needed")
                return True

            # Create new entries
            new_ucc = self.missing_entries[['cell', 'modyy']].copy()
            new_ucc['ucc'] = "PENDING"
            new_ucc['description'] = "PENDING DESCRIPTION"

            # Read existing UCC data
            ucc_path = self.ucc_entry.get()
            ucc_path_obj = Path(ucc_path)
            
            if not ucc_path_obj.exists():
                raise FileNotFoundError(f"UCC file not found: {ucc_path}")
            
            # Read with appropriate engine/encoding
            if ucc_path.lower().endswith(('.xls', '.xlsx')):
                existing_ucc = pd.read_excel(ucc_path, dtype=str, engine='openpyxl')
            else:
                try:
                    existing_ucc = pd.read_csv(ucc_path, dtype=str, encoding='utf-8')
                except UnicodeDecodeError:
                    existing_ucc = pd.read_csv(ucc_path, dtype=str, encoding='latin-1')
            
            # Standardize columns
            existing_ucc = self.standardize_columns(existing_ucc)
            existing_ucc.columns = [col.strip().lower() for col in existing_ucc.columns]

            # Normalize UCC column name
            ucc_col = next((col for col in existing_ucc.columns if 'ucc' in col.lower()), None)
            if ucc_col:
                existing_ucc.rename(columns={ucc_col: 'ucc'}, inplace=True)
            else:
                existing_ucc['ucc'] = ''

            # Create timestamped backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{ucc_path}.{timestamp}.bak"
            shutil.copy2(ucc_path, backup_path)
            self.log(f"Created timestamped UCC backup: {backup_path}")

            # Merge and deduplicate
            updated_ucc = pd.concat([existing_ucc, new_ucc])
            
            # Add missing columns if needed
            for col in ['cell', 'modyy', 'ucc', 'description']:
                if col not in updated_ucc:
                    updated_ucc[col] = ''
                    
            # Deduplicate keeping last occurrence (new entries take priority)
            updated_ucc = updated_ucc.drop_duplicates(
                subset=['cell', 'modyy'], 
                keep='last'
            )
            
            # Save in original format
            if ucc_path.lower().endswith(('.xls', '.xlsx')):
                updated_ucc.to_excel(ucc_path, index=False, engine='openpyxl')
            else:
                # Maintain original encoding
                updated_ucc.to_csv(ucc_path, index=False, encoding='utf-8')
            
            self.log(f"Updated UCC master with {len(new_ucc)} new entries")
            self.status(f"UCC master updated with {len(new_ucc)} new entries")
            self.ucc_updated = True
            return True

        except Exception as e:
            import traceback
            error_msg = f"UCC update failed: {str(e)}\n{traceback.format_exc()}"
            self.log(error_msg)
            self.status("UCC update failed")
            return False
        

    def step_ucc_merge(self):
        try:
            input_path = Path(self.output_entry.get()) / "ucc_checked.csv"
            ucc_path = Path(self.ucc_entry.get())
            output_path = Path(self.output_entry.get()) / "merged_with_ucc.csv"

            # 1. Validate input files
            if not input_path.exists():
                raise FileNotFoundError(f"Input file not found: {input_path}")
            if not ucc_path.exists():
                raise FileNotFoundError(f"UCC file not found: {ucc_path}")

            # 2. Read data
            current_df = pd.read_csv(input_path, dtype=str)
            current_df.columns = [col.strip().lower() for col in current_df.columns]

            # 3. Read UCC file
            if str(ucc_path).lower().endswith(('.xls', '.xlsx')):
                ucc_df = pd.read_excel(ucc_path, sheet_name=0, dtype=str)
                ucc_df = ucc_df.dropna(how='all').dropna(axis=1, how='all')
            else:
                try:
                    ucc_df = pd.read_csv(ucc_path, dtype=str, encoding='utf-8', on_bad_lines='warn')
                except UnicodeDecodeError:
                    ucc_df = pd.read_csv(ucc_path, dtype=str, encoding='latin-1', on_bad_lines='warn')

            # 4. Standardize UCC columns
            ucc_df = self.standardize_columns(ucc_df)
            ucc_df.columns = [col.strip().lower() for col in ucc_df.columns]
            
            # 4.1 DEBUG: Log UCC columns
            self.log(f"UCC file columns: {ucc_df.columns.tolist()}")
            self.log(f"Sample UCC data: {ucc_df[['cell', 'year', 'ucc']].head(3).to_dict('records')}")

            # 5. Map UCC file columns to required names
            column_map = {
                'year': 'modyy',
                'cell': 'cell',
                'ucc': 'ucc'
            }
            for src_col, dest_col in column_map.items():
                if src_col in ucc_df.columns and dest_col not in ucc_df.columns:
                    ucc_df[dest_col] = ucc_df[src_col]

            # 6. Ensure required columns
            required_columns = ['cell', 'modyy', 'ucc']
            for col in required_columns:
                if col not in ucc_df.columns:
                    self.log(f"WARNING: UCC file missing '{col}'. Using placeholder")
                    ucc_df[col] = ''

            # 7. Clean and normalize key columns
            for df in [current_df, ucc_df]:
                df['cell'] = df['cell'].astype(str).str.strip().str.upper()
                df['modyy'] = df['modyy'].astype(str).str.strip()
                
                # Convert year to 4-digit format
                df['modyy'] = df['modyy'].apply(
                    lambda x: 
                        str(int(float(x))) if '.' in x else  # Handle 2010.0
                        '20' + x.zfill(2) if len(x) == 2 else  # Handle 14 → 2014
                        x
                )

            # Before UCC merge, add:
            current_df['cell'] = current_df['cell'].astype(str).str.strip()

            # 8. DEBUG: Log key samples
            unique_combos = current_df[['cell','modyy']].drop_duplicates()
            sample_size = min(5, len(unique_combos))
            if sample_size > 0:
                sample_data = unique_combos.head(sample_size).values.tolist()
                self.log(f"Vehicle keys sample: {sample_data}")
            else:
                self.log("No vehicle keys available for sampling")
            self.log(f"UCC keys sample: {ucc_df[['cell','modyy']].drop_duplicates().sample(5).values.tolist()}")

            # 9. Perform merge with explicit column handling
            # Rename UCC column to avoid conflicts
            ucc_df = ucc_df.rename(columns={'ucc': 'ucc_value'})
            
            merged_df = current_df.merge(
                ucc_df[['cell', 'modyy', 'ucc_value']],
                on=['cell', 'modyy'],
                how='left'
            )
            
            # 10. Create final UCC column
            merged_df['ucc'] = merged_df['ucc_value'].fillna('MISSING_UCC')
            
            # 11. Remove temporary column
            merged_df = merged_df.drop(columns=['ucc_value'])
            
            # 12. Log UCC samples
            self.log(f"Pre-merge UCC sample: {ucc_df['ucc_value'].head(3).tolist()}")
            self.log(f"Post-merge UCC sample: {merged_df['ucc'].head(3).tolist()}")
            
            # 13. Log merge results
            match_count = (merged_df['ucc'] != 'MISSING_UCC').sum()
            total_records = len(merged_df)
            self.log(f"UCC merge results: {match_count} matched, {total_records - match_count} missing")
            
            # 14. Clean UCC values
            merged_df['ucc'] = merged_df['ucc'].astype(str).str.strip()
            merged_df['ucc'] = merged_df['ucc'].replace({'nan': '', '': 'MISSING_UCC'})

            # 15. Handle missing UCC codes
            missing_mask = merged_df['ucc'] == 'MISSING_UCC'
            missing_count = missing_mask.sum()
            if missing_count > 0:
                missing_combos = merged_df.loc[missing_mask, ['cell', 'modyy']].drop_duplicates()
                missing_path = Path(self.output_entry.get()) / "missing_ucc_final.csv"
                missing_combos.to_csv(missing_path, index=False)
                self.log(f"WARNING: {missing_count} records with missing UCC. Saved {len(missing_combos)} unique combinations")

            # In step_ucc_merge(), before saving:
            self.log("Formatting UCC codes for Excel compatibility...")

            def format_ucc(x):
                """Format 21-digit UCC codes for Excel"""
                if isinstance(x, str) and x.isdigit() and len(x) == 21:
                    return f'="{x}"'  # Excel-safe format
                return x

            merged_df['ucc'] = merged_df['ucc'].apply(format_ucc)

            # 16. Save output
            merged_df.to_csv(output_path, index=False)
            self.log(f"Successfully merged UCC codes to {len(merged_df)} records")
            return True

        except Exception as e:
            import traceback
            error_msg = f"UCC merge failed: {str(e)}\n{traceback.format_exc()}"
            self.log(error_msg)
            return False


    def step_electric_merge(self):
        try:
            input_path = Path(self.output_entry.get()) / "merged_with_ucc.csv"
            output_path = Path(self.output_entry.get()) / "merged_electric.csv"
            
            df = pd.read_csv(input_path, dtype=str)
            
            # Add electric flag column
            df['electric'] = '0'  # Default to non-electric
            
            # Simple electric detection (customize as needed)
            electric_keywords = ['electric', 'ev', 'phev', 'hybrid']
            for col in ['model', 'trimline', 'description']:
                if col in df.columns:
                    for keyword in electric_keywords:
                        mask = df[col].str.contains(keyword, case=False, na=False)
                        df.loc[mask, 'electric'] = '1'
            
            df.to_csv(output_path, index=False)
            self.log(f"Electric Merge: Flagged {len(df[df['electric']=='1'])} electric vehicles")
            return True
            
        except Exception as e:
            self.log(f"Electric merge failed: {str(e)}")
            self.status("Electric merge failed")
            return False




    def step_desc_merge(self):
        try:
            input_path = Path(self.output_entry.get()) / "merged_electric.csv"
            output_path = Path(self.output_entry.get()) / "merged_with_desc.csv"
            desc_path = self.desc_entry.get()
            
            # Load main data
            df = pd.read_csv(input_path, dtype=str)
            df = self.standardize_columns(df)
            self.log(f"Columns in merged_electric.csv: {df.columns.tolist()}")

            # Standardize cell column
            df['cell'] = self.remove_decimal_zero(df['cell'].astype(str).str.strip().str.upper())

            # Load and standardize description file
            if desc_path.endswith('.xlsx'):
                desc_df = pd.read_excel(desc_path, dtype=str)
            else:
                desc_df = pd.read_csv(desc_path, dtype=str, encoding='utf-8')
            desc_df.columns = [col.strip().lower() for col in desc_df.columns]

            # Find and rename description column
            desc_col = next((col for col in desc_df.columns if 'mailing description' in col or 'description' in col), None)
            if not desc_col:
                raise ValueError(f"No description column found. Columns: {desc_df.columns.tolist()}")
            desc_df = desc_df.rename(columns={desc_col: 'description_long'})
            desc_df['description_long'] = desc_df['description_long'].str.strip().fillna('')

            # Standardize cell in description file
            desc_df['cell'] = self.remove_decimal_zero(desc_df['cell'].astype(str).str.strip().str.upper())

            # Remove duplicate cell entries (keep first)
            desc_df = desc_df.drop_duplicates(subset=['cell'], keep='first')

            # Merge using only 'cell'
            merged_df = pd.merge(
                df,
                desc_df[['cell', 'description_long']],
                on='cell',
                how='left'
            )

            # Handle missing description_long
            missing_desc = merged_df['description_long'].isna().sum()
            self.tracker.set('missing_descriptions', missing_desc)
            merged_df['description_long'] = merged_df['description_long'].fillna('MISSING DESCRIPTION')

            # Extract and save missing cell codes
            missing_combos = merged_df[merged_df['description_long'] == 'MISSING DESCRIPTION'][['cell']].drop_duplicates()
            if not missing_combos.empty:
                missing_path = Path(self.output_entry.get()) / "missing_desc_combinations.csv"
                missing_combos.to_csv(missing_path, index=False)
                self.log(f"Saved {len(missing_combos)} missing description cell codes to {missing_path}")

            # Final output
            merged_df.to_csv(output_path, index=False)
            self.log(f"Description merge: Added long descriptions to {len(merged_df) - missing_desc}/{len(df)} records ({missing_desc} missing)")
            return True

        except Exception as e:
            import traceback
            self.log(f"Description merge error: {str(e)}\n{traceback.format_exc()}")
            self.status("Description merge failed")
            return False



    def step_assign_sequence(self):
        try:
            input_path = Path(self.output_entry.get()) / "merged_with_desc.csv"
            output_path = Path(self.output_entry.get()) / "sequenced_data.csv"
            
            df = pd.read_csv(input_path, dtype=str)
            df = self.standardize_columns(df)

            if df.empty:
                self.log("No data to assign sequence numbers.")
                return False

            if 'purchase_date' not in df.columns:
                self.log("Missing 'purchase_date' column in input data.")
                return False

            # Use the enhanced parser
            df['purchase_date'] = df['purchase_date'].apply(self.parse_purchase_date)
            
            # Extract month/year from parsed datetime
            df['MO Purch'] = df['purchase_date'].dt.strftime('%m').replace('NaT', '').fillna('')
            df['YR Purch'] = df['purchase_date'].dt.strftime('%Y').replace('NaT', '').fillna('')

            # Add CHANNEL column: "email" if email present and non-empty, else "mail"
            def determine_channel(row):
                email = row.get('email') or row.get('EMAIL')
                if pd.isna(email):
                    return 'mail'
                email = str(email).strip().lower()
                return 'email' if email and email != 'nan' else 'mail'

            # Support both 'email' and 'EMAIL' column casing
            if 'email' in df.columns or 'EMAIL' in df.columns:
                df['channel'] = df.apply(determine_channel, axis=1)
            else:
                df['channel'] = 'mail'

            # Sequence assignment logic
            try:
                start_seq = int(self.sequence_entry.get())
            except ValueError:
                self.log("Sequence assignment failed: Sequence # must be an integer.")
                return False
            
            df['seq8'] = list(range(start_seq, start_seq + len(df)))
            last_seq = start_seq + len(df) - 1 if len(df) > 0 else start_seq

            # Handle invalid dates
            invalid_dates = df[(df['MO Purch'] == '') | (df['YR Purch'] == '')]
            if not invalid_dates.empty:
                invalid_path = Path(self.output_entry.get()) / "invalid_dates.csv"
                invalid_dates.to_csv(invalid_path, index=False)
                self.log(f"Found {len(invalid_dates)} records with invalid dates - saved to {invalid_path.name}")
            
            df.to_csv(output_path, index=False)
            self.log(f"Assigned sequence numbers from {start_seq} to {last_seq} ({len(df)} records), channel assigned.")

            return True
        except Exception as e:
            self.log(f"Sequence assignment failed: {str(e)}")
            return False




    def parse_purchase_date(self, date_str):
        """Parse OEM-specific date formats into datetime objects"""
        try:
            if pd.isna(date_str) or str(date_str).strip() in ['', 'nan', 'NaT']:
                return pd.NaT

            s = str(date_str).strip()
            manufacturer = self.manufacturer_combo.get().lower().strip()
            
            # Mercedes-specific parsing
            if manufacturer == 'mercedes':
                # Format: "2025-05-16-18.19.43.000000000"
                if len(s) >= 19:
                    # Extract components using fixed positions
                    year = s[0:4]
                    month = s[5:7]
                    day = s[8:10]
                    hour = s[11:13]
                    minute = s[14:16]
                    second = s[17:19]
                    # Build ISO format string
                    iso_str = f"{year}-{month}-{day} {hour}:{minute}:{second}"
                    return datetime.strptime(iso_str, "%Y-%m-%d %H:%M:%S")
                return pd.NaT
            
            # BMW format (mm/dd/yyyy)
            if '/' in str(date_str):
                parts = str(date_str).split('/')
                if len(parts) == 3:
                    return datetime.strptime(date_str, "%m/%d/%Y")

            # JLR format (YYYYMMDD)
            if len(str(date_str)) == 8 and str(date_str).isdigit():
                return datetime.strptime(str(date_str), "%Y%m%d")
            
            # BMW/Honda/Nissan format
            elif '/' in str(date_str):
                return pd.to_datetime(date_str, errors='coerce')
            
            # Fallback
            else:
                return pd.to_datetime(date_str, errors='coerce')
        except Exception:
            return pd.NaT


    def step_combine_rejections(self):
        try:
            output_dir = Path(self.output_entry.get())
            rejection_files = [
                "name_removed_blanks.csv",
                "email_rejected.csv",
                "geo_rejected.csv", 
                "business_rejected.csv",
                "vin_rejected.csv",
                "sales_type_rejected.csv",
                "duplicates_rejected.csv",   # <--- Add comma here!
                "filtered_out_duplicate_address.csv"
            ]

            all_rejections = []
            for fname in rejection_files:
                fpath = output_dir / fname
                if fpath.exists() and os.path.getsize(fpath) > 0:  # Skip empty files
                    df = pd.read_csv(fpath)
                    df = self.standardize_columns(df)
                    if not df.empty:
                        # Cleaner source name
                        df['rejection_source'] = fname.replace('.csv', '')
                        all_rejections.append(df)
            
            if all_rejections:
                master_rejections = pd.concat(all_rejections, ignore_index=True)
                master_rejections.to_csv(output_dir / "master_rejection_report.csv", index=False)
                self.log(f"Combined {len(all_rejections)} rejection files into master report")
                self.status(f"Combined {len(all_rejections)} rejection files")
            else:
                self.log("No rejection files found or all are empty.")
                self.status("No rejections to combine.")
            return True
        except Exception as e:
            self.log(f"Rejection combining failed: {str(e)}")
            self.status("Rejection combining failed")
            return False



    def step_cellcode_reporting(self):
        try:
            import pandas as pd
            input_path = Path(self.output_entry.get()) / "sequenced_data.csv"
            output_path = Path(self.output_entry.get()) / "cellcode_report.csv"

            # --- Load data ---
            cellcode_df = pd.read_csv(input_path, dtype=str)
            self.log(f"Step cellcode_reporting: Loaded {len(cellcode_df)} rows from sequenced_data.csv")
            cellcode_df = self.standardize_columns(cellcode_df)
            cellcode_df = cellcode_df.loc[:, ~cellcode_df.columns.duplicated()]

            # --- Ensure modyy ---
            year_cols = ['modyy', 'model_year', 'modelyear', 'year']
            modyy_col = next((col for col in year_cols if col in cellcode_df.columns), None)
            if modyy_col and modyy_col != 'modyy':
                self.log(f"Renamed '{modyy_col}' to 'modyy' for reporting")
                cellcode_df.rename(columns={modyy_col: 'modyy'}, inplace=True)
            elif 'modyy' not in cellcode_df.columns:
                self.log("WARNING: 'modyy' column missing. Creating empty column.")
                cellcode_df['modyy'] = ''
            cellcode_df['modyy'] = cellcode_df['modyy'].astype(str).str.strip()

            # --- Find description column ---
            possible_desc_cols = ['ldesc', 'desc', 'description', 'mailing description', 'mail_desc', 'long_desc']
            desc_col = next((col for col in cellcode_df.columns if col in possible_desc_cols), None)
            if not desc_col:
                desc_col = 'description_fallback'
                cellcode_df[desc_col] = 'UNKNOWN'
                self.log(f"No description column found. Using fallback: {desc_col}")

            # --- Ensure CHANNEL column exists ---
            if 'channel' not in cellcode_df.columns:
                self.log("WARNING: 'channel' column missing; assigning 'unknown'")
                cellcode_df['channel'] = 'unknown'
            else:
                cellcode_df['channel'] = cellcode_df['channel'].fillna('unknown').str.lower().str.strip()

            # --- Build composite pivot column: Year + Channel ---
            cellcode_df['pivcol'] = (
                'Year ' + cellcode_df['modyy'].astype(str).str.strip() +
                ' ' + cellcode_df['channel'].str.title()
            )

            # --- Separate coded (nonblank/non-default) and uncoded (blank/'default') ---
            # Treat records with cell in [empty, "default", "DEFAULT", etc] as uncoded
            cell_col_upper = cellcode_df['cell'].str.upper().fillna('') if 'cell' in cellcode_df.columns else pd.Series('', index=cellcode_df.index)
            def is_coded(x):
                if pd.isna(x):
                    return False
                val = str(x).strip().upper()
                return val not in ['', 'DEFAULT', 'UNCODED', 'MISSING']

            mask_coded = cell_col_upper.apply(is_coded)
            df_coded = cellcode_df[mask_coded].copy()
            df_uncoded = cellcode_df[~mask_coded].copy()


            # --- Standard pivot for coded ---
            if not df_coded.empty:
                df_coded['cell'] = df_coded['cell'].str.strip()
                df_coded[desc_col] = df_coded[desc_col].str.strip()
                coded_pivot = pd.pivot_table(
                    df_coded,
                    index=['cell', desc_col],
                    columns='pivcol',
                    values='vin',
                    aggfunc='count',
                    fill_value=0
                )
                coded_pivot['Total'] = coded_pivot.sum(axis=1)
                coded_pivot = coded_pivot.reset_index()
            else:
                coded_pivot = pd.DataFrame()

            # --- Pivot for uncoded (if any) ---
            if not df_uncoded.empty:
                df_uncoded = df_uncoded.copy()
                df_uncoded['cell'] = 'UNCODED'
                df_uncoded[desc_col] = 'Uncoded – No Cell Assigned'
                uncoded_pivot = pd.pivot_table(
                    df_uncoded,
                    index=['cell', desc_col],
                    columns='pivcol',
                    values='vin',
                    aggfunc='count',
                    fill_value=0
                )
                uncoded_pivot['Total'] = uncoded_pivot.sum(axis=1)
                uncoded_pivot = uncoded_pivot.reset_index()
            else:
                uncoded_pivot = pd.DataFrame()

            # --- Combine and save ---
            if not coded_pivot.empty and not uncoded_pivot.empty:
                final_pivot = pd.concat([coded_pivot, uncoded_pivot], ignore_index=True)
            elif not coded_pivot.empty:
                final_pivot = coded_pivot
            elif not uncoded_pivot.empty:
                final_pivot = uncoded_pivot
            else:
                final_pivot = pd.DataFrame(columns=['cell', desc_col])

            final_pivot.to_csv(output_path, index=False)
            self.log(f"Step cellcode_reporting: Saved detailed year+channel split to {output_path}. Included uncoded row if any.")

            return True

        except Exception as e:
            import traceback
            self.log(f"Cellcode reporting failed: {str(e)}\n{traceback.format_exc()}")
            return False




    from datetime import datetime
    import pandas as pd
    import yaml
    from pathlib import Path

    def step_final_outputs(self):
        try:
            input_path = Path(self.output_entry.get()) / "sequenced_data.csv"
            output_dir = Path(self.output_entry.get())
            
            # 1. Read data and remove duplicate columns
            df = pd.read_csv(input_path, dtype=str)
            df = df.loc[:, ~df.columns.duplicated()]
            
            # 2. Safe domain extraction from email
            def safe_domain_extraction(email):
                if pd.isna(email) or not isinstance(email, str):
                    return 'unknown'
                return email.split('@')[-1] if '@' in email else 'unknown'
            df['DOMAIN'] = df['email'].apply(safe_domain_extraction)
            
            # 3. Rename bodycode to MFG MODEL CODE (only once)
            if 'bodycode' in df.columns:
                df = df.rename(columns={'bodycode': 'MFG MODEL CODE'})
            else:
                df['MFG MODEL CODE'] = ''
            
            # 4. Add Brand and Description from YAML
            manufacturer = self.manufacturer_combo.get().lower().strip()

            vin_map_path = self._get_vin_mapping_path(manufacturer, self.vin_mapping_dir)
            with open(vin_map_path, encoding="utf-8") as f:
                vin_mapping = yaml.safe_load(f)
            vin_rules = vin_mapping.get('vin_mappings', [])
            
            def get_yaml_value(row, key):
                """Robust YAML value extractor with normalization"""
                vin13 = str(row.get('vin13', '')).strip().upper()
                vin48 = str(row.get('vin48', '')).strip().upper()
                vin10 = str(row.get('vin10', '')).strip().upper()
                
                for rule in vin_rules:
                    rule_vin13 = str(rule.get('vin13', '')).strip().upper()
                    rule_vin48 = str(rule.get('vin48', '')).strip().upper()
                    rule_vin10 = str(rule.get('vin10', '')).strip().upper()
                    
                    if (rule_vin13 == vin13 and 
                        rule_vin48 == vin48 and 
                        rule_vin10 == vin10):
                        return rule.get(key, '')
                return ''
            
            # Add Brand and Description columns
            df['Brand'] = df.apply(lambda row: get_yaml_value(row, 'brand'), axis=1)
            df['Description'] = df.apply(lambda row: get_yaml_value(row, 'desc'), axis=1)

            # Remove duplicate or conflicting columns (case-insensitive)
            for col in ['brand', 'Brand', 'BRAND']:
                if col in df.columns and col != 'Brand':
                    df = df.drop(columns=[col])

            # Ensure only 'Brand' exists
            if 'Brand' not in df.columns:
                df['Brand'] = ''

            for col in ['desc', 'Description', 'DESCRIPTION']:
                if col in df.columns and col != 'Description':
                    df = df.drop(columns=[col])
            if 'Description' not in df.columns:
                df['Description'] = ''


            # 5. Parse purchase_date to extract YR Purch and MO Purch
            # COMMENTED OUT THIS FULL SECTION - WE ARE UPDATING PARSE_PURCHASE_DATE AND
            # STEP_ASSIGN_SEQUENCE - IF IT WORKS WELL WE WILL DELETE THIS COMMENTED SECTION.
            #if 'purchase_date' in df.columns:

                # Extract year (first 4 characters always work)
                #df['YR Purch'] = df['purchase_date'].str[:4]
                
                # Extract month with Mercedes-safe method
                #def extract_month(date_str):
                    #if pd.isna(date_str) or not isinstance(date_str, str):
                        #return ''
                    #parts = date_str.split('-')
                    #if len(parts) >= 2:
                        # Month is between first and second hyphen: "2025-05-16-..."
                        #month_part = parts[1] if len(parts) > 1 else ''
                        #return month_part[:2]  # Take first 2 chars of month part
                    #return ''
                
                #df['MO Purch'] = df['purchase_date'].apply(extract_month).str.zfill(2)
            #else:
                #df['YR Purch'] = ''
                #df['MO Purch'] = ''


            # 6. Consolidate purchase date columns
            # Handle case variations from previous steps
            if 'MO Purch' in df.columns and 'MO PURCH' in df.columns:
                df['MO Purch'] = df['MO Purch'].combine_first(df['MO PURCH'])
                df = df.drop(columns=['MO PURCH'])
            if 'YR Purch' in df.columns and 'YR PURCH' in df.columns:
                df['YR Purch'] = df['YR Purch'].combine_first(df['YR PURCH'])
                df = df.drop(columns=['YR PURCH'])
            
            # 7. Ensure consistent column names
            df = df.rename(columns={
                'MO Purch': 'MO Purch',
                'YR Purch': 'YR Purch'
            })
            
            # 8. Validate critical columns
            required_columns = ['Brand', 'cell', 'email', 'ucc', 'seq8']
            for col in required_columns:
                if col not in df.columns:
                    df[col] = 'MISSING'
                    self.log(f"CRITICAL: Added missing {col} column")
            
            # 9. Add panel metadata columns
            df['ROTATION'] = 'A'
            df['VERSION'] = '1.0'
            df['FILE SOURCE'] = Path(self.input_entry.get()).name
            df['EXTERNALREFERENCE'] = df['seq8'] if 'seq8' in df.columns else ''
            df['PHONE'] = ''

            # Rename 'email' to uppercase if it exists
            if 'email' in df.columns:
                df.rename(columns={'email': 'EMAIL'}, inplace=True)
            elif 'EMAIL' not in df.columns:
                df['EMAIL'] = ''  # Fallback if missing entirely

            # 10. Save final output CSV
            month = self.month_entry.get().strip()
            year = self.year_entry.get().strip()
            final_path = output_dir / f"{manufacturer}_{month}_{year}_FINAL_OUTPUT.csv"
            df.to_csv(final_path, index=False)
            
            # 11. Log success
            self.log(f"Final output saved: {final_path} with {len(df)} records")
            self.log(f"Final output columns: {list(df.columns)}")
            
            return True
        
        except Exception as e:
            import traceback
            self.log(f"Final output failed: {str(e)}\n{traceback.format_exc()}")
            self.status("Final output failed")
            return False







    def step_generate_report(self):
        try:
            metrics = self.tracker.get_metrics()
            
            # Calculate derived metrics
            total_geo_rejections = metrics.get('bad_state_codes', 0) + metrics.get('bad_zip_state', 0)
            total_deduped = metrics.get('historical_deduped', 0) + metrics.get('truecar_deduped', 0)
            
            # Get final count from the last processing step
            final_output_path = Path(self.output_entry.get()) / "final_output.csv"
            if final_output_path.exists():
                final_count = sum(1 for _ in open(final_output_path)) - 1  # Subtract header
            else:
                final_count = metrics.get('dedup_remaining', 0)  # Use dedup_remaining as fallback

            report_content = [
                "# Data Processing Report",
                f"**Manufacturer:** {self.manufacturer_combo.get()}",
                f"**Processing Date:** {self.month_entry.get()}/{self.year_entry.get()}",
                f"**Sequence #:** {self.sequence_entry.get()}",
                "\n## Processing Metrics",
                f"- Total Records: {metrics.get('initial_count', 0):,}",
                f"- Invalid Sales Type: {metrics.get('invalid_sales_type', 0):,}",
                f"- Invalid Model Year: {metrics.get('invalid_model_year', 0):,}",
                f"- Invalid Emails: {metrics.get('invalid_emails', 0):,}",
                f"- Blank Zip: {metrics.get('blank_zip', 0):,}",
                f"- Blank Name: {metrics.get('blank_name', 0):,}",
                f"- Duplicate Addresses: {metrics.get('duplicate_addresses', 0):,}",
                f"- Bad State Codes: {metrics.get('bad_state_codes', 0):,}",
                f"- State/ZIP Mismatches: {metrics.get('bad_zip_state', 0):,}",
                f"- Business Exclusions: {metrics.get('business_exclusions', 0):,}",
                f"- Duplicate VINs Removed: {metrics.get('duplicate_vins', 0):,}",
                f"- VINs Rejected: {metrics.get('vin_rejected', 0):,}",
                f"- VINs Missing Cell: {metrics.get('vin_missing_cell', 0):,}",
                f"- VINs Missing ModelID: {metrics.get('vin_missing_modelid', 0):,}",
                f"- Deduplication Initial Count: {metrics.get('dedup_initial', 0):,}",
                f"- Deduplication Removed: {metrics.get('dedup_removed', 0):,}",
                f"- Deduplication Remaining: {metrics.get('dedup_remaining', 0):,}",
                f"- Historical Duplicates Removed: {metrics.get('historical_deduped', 0):,}",
                f"- TrueCar Duplicates Removed: {metrics.get('truecar_deduped', 0):,}",
                f"- Total Duplicates Removed: {total_deduped:,}",
                f"- Missing Descriptions: {metrics.get('missing_descriptions', 0):,}",
                f"- Blank Sales Type: {metrics.get('blank_sales_type', 0):,}",
                f"- Final Valid Records: {final_count:,}"
            ]
            
            output_dir = Path(self.output_entry.get())
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / "final_report.md"
            
            with open(output_path, 'w') as f:
                f.write('\n'.join(report_content))
            
            self.log(f"Generated final report at {output_path}")
            return True
            
        except Exception as e:
            self.log(f"Report generation failed: {str(e)}\n{traceback.format_exc()}")
            return False


    def threaded_excel_export(self, df, path):
        threading.Thread(target=self.export_task, args=(df, path), daemon=True).start()

    # Add this class method
    def export_task(self, df, path):
        try:
            from openpyxl.styles import Font, Alignment
            df.to_excel(path, index=False)
            wb = load_workbook(path)
            ws = wb.active
            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)
            # Center specific columns
            for col_name in ['ROTATION', 'VERSION', 'MOD YR']:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
            wb.save(path)
        except Exception as e:
            self.log(f"Excel export error: {str(e)}")


        threading.Thread(target=export_task, daemon=True).start()


    def export_to_excel_with_formatting(self, df, path):
        """Save DataFrame to Excel with formatting, integrated with app logging/UI"""
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        try:
            # Use ExcelWriter context manager for better resource handling
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data')
                workbook = writer.book
                worksheet = writer.sheets['Data']

                # Format headers during initial write (no need to reload)
                header_font = Font(bold=True, color="FFFFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = str(col_name).upper()
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Center specified columns
                center_cols = ['ROTATION', 'VERSION', 'MOD YR']
                for col_name in center_cols:
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        for row in worksheet.iter_rows(
                            min_row=2, 
                            max_row=worksheet.max_row,
                            min_col=col_idx,
                            max_col=col_idx
                        ):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center')

                # Auto-adjust column widths
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells) + 2
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length

            self.log(f"Excel file formatted and saved to {path}")
            return True

        except PermissionError as pe:
            error_msg = f"Permission denied: {pe}\nIs the file open in Excel?"
            self.log(error_msg)
            messagebox.showerror(
                "Save Error", 
                f"Cannot save to {path}\nClose the file in Excel and try again.",
                parent=self.root
            )
            return False
            
        except Exception as e:
            error_msg = f"Excel export failed: {str(e)}"
            self.log(error_msg)
            messagebox.showerror(
                "Export Error", 
                error_msg,
                parent=self.root
            )
            return False



    def step_panelization(self):
        try:
            manufacturer = self.manufacturer_combo.get().replace(' ', '_')
            month = self.month_entry.get().strip()
            year = self.year_entry.get().strip()
            base_name = f"{manufacturer}_{month}_{year}"
            final_output_path = Path(self.output_entry.get()) / f"{base_name}_FINAL_OUTPUT.csv"

            if not final_output_path.exists():
                self.log(f"Final output file not found: {final_output_path}")
                return False

            df = pd.read_csv(final_output_path, dtype=str)
            self.log(f"Loaded final output with {len(df)} records")

           # Convert all column names to uppercase
            df.columns = df.columns.str.upper()
            self.log(f"Step Panelization: Uppercased columns: {df.columns.tolist()}")


            # Define column mapping (output column → source column)
            column_mapping = {
                #'ZIP': 'ZIP_CODE',
                'SVI CELL': 'CELL',
                'MOD YR': 'MODYY',
                'DESCRIPTION': 'DESCRIPTION',  # Changed from 'DESC'
                'SEQUENCE #(8)': 'SEQ8',
                'MFG MODEL CODE': 'MFG MODEL CODE',  # Direct mapping
                'ADDRESS': 'ADDRESS_LINE1',  # CRITICAL ADDITION
                'BRAND': 'BRAND'  # Added
            }
            
            # --- Define columns for email and mail files ---
            email_columns = [
                'NAME', 'ADDRESS', 'CITY', 'STATE', 'ZIP', 'ROTATION', 'VERSION',
                'FILE SOURCE', 'SVI CELL', 'MOD YR', 'DESCRIPTION', 'VIN', 'PURCH/LSE',
                'MO PURCH', 'YR PURCH', 'MAIL DATE (MMDD)', 'BRAND', 'SEQUENCE #(8)',
                'EMAIL', 'MFG MODEL CODE', 'EXTRA1', 'UCC', 'EXTERNALREFERENCE',
                'ELECTRIC', 'SOS', 'DOMAIN'
            ]
            
            mail_columns = [
                'NAME', 'ADDRESS', 'CITY', 'STATE', 'ZIP', 'ROTATION', 'VERSION',
                'FILE SOURCE', 'SVI CELL', 'MOD YR', 'DESCRIPTION', 'VIN', 'PURCH/LSE',
                'MO PURCH', 'YR PURCH', 'MAIL DATE (MMDD)', 'BRAND', 'SEQUENCE #(8)',
                'MFG MODEL CODE', 'EXTRA1', 'UCC', 'EXTERNALREFERENCE',
                'ELECTRIC', 'SOS', 'DOMAIN'
            ]

            self.log(f"Total records loaded for split: {len(df)}")
            self.log(f"EMAIL sample values: {df['EMAIL'].head(5).tolist()}")

            # --- Split email vs mail ---
            if 'EMAIL' not in df.columns:
                self.log("ERROR: EMAIL column missing before panel split!")
                return False

            email_mask = df['EMAIL'].fillna('').str.strip().str.lower().replace('nan', '').astype(bool)
            email_df = df[email_mask].copy()
            mail_df = df[~email_mask].copy()

            self.tracker.set('email_count', len(email_df))
            self.tracker.set('mail_count', len(mail_df))
            self.log(f"Split complete: {len(email_df)} email records, {len(mail_df)} mail records.")


            # --- Create output DataFrames with mapped data ---
            def create_output_df(source_df, columns):
                output_df = pd.DataFrame()
                for col in columns:
                    # Handle special cases
                    if col == 'PURCH/LSE':
                        output_df[col] = '1'  # Hardcoded value
                    # Map from source columns
                    elif col in column_mapping:
                        source_col = column_mapping[col]
                        # Use .get() with default value if column doesn't exist
                        output_df[col] = source_df.get(source_col, '')
                    # Copy directly if exists
                    elif col in source_df.columns:
                        output_df[col] = source_df[col]
                    # Add empty column if missing
                    else:
                        output_df[col] = ''
                return output_df

            # Create the output DataFrames
            email_output_df = create_output_df(email_df, email_columns)
            mail_output_df = create_output_df(mail_df, mail_columns)

            total = len(email_output_df) + len(mail_output_df)
            self.tracker.set('total_panel_records', total)
            self.log(f"Total panel records processed: {total}")

            # Create channel column
            email_output_df['CHANNEL'] = 'email'
            mail_output_df['CHANNEL'] = 'mail'

            self.log(f"Email records with channel 'email': {email_output_df['CHANNEL'].eq('email').sum()}")
            self.log(f"Mail records with channel 'mail': {mail_output_df['CHANNEL'].eq('mail').sum()}")

            # --- Save email and mail files ---
            email_output_path = Path(self.output_entry.get()) / f"{base_name}_email.csv"
            mail_output_path = Path(self.output_entry.get()) / f"{base_name}_mail.csv"
            email_output_df.to_csv(email_output_path, index=False)
            mail_output_df.to_csv(mail_output_path, index=False)
            self.log(f"Saved email file ({len(email_output_df)} records) to: {email_output_path}")
            self.log(f"Saved mail file ({len(mail_output_df)} records) to: {mail_output_path}")

            # --- Panelization for email records ---
            panel_dir = Path(self.output_entry.get()) / "panels"
            panel_dir.mkdir(parents=True, exist_ok=True)

            if not email_output_df.empty:
                email_output_df['domain'] = email_output_df['EMAIL'].str.split('@').str[-1].str.lower()
                panel_types = {
                    'gmail': ['gmail.com', 'googlemail.com'],
                    'yahoo_aol': ['yahoo.com', 'aol.com', 'ymail.com', 'rocketmail.com'],
                    'other': []  # Everything else
                }
                for panel_name, domains in panel_types.items():
                    if panel_name == 'other':
                        panel_df = email_output_df[~email_output_df['domain'].isin(
                            panel_types['gmail'] + panel_types['yahoo_aol']
                        )]
                    else:
                        panel_df = email_output_df[email_output_df['domain'].isin(domains)]
                    if panel_df.empty:
                        self.log(f"No records found for {panel_name} panel")
                        continue
                    chunk_size = 4500
                    num_chunks = (len(panel_df) // chunk_size) + 1
                    for chunk_idx in range(num_chunks):
                        start_idx = chunk_idx * chunk_size
                        end_idx = min((chunk_idx + 1) * chunk_size, len(panel_df))
                        chunk_df = panel_df.iloc[start_idx:end_idx].copy()
                        if chunk_df.empty:
                            continue
                        panel_path = panel_dir / (
                            f"{base_name}_{panel_name}_panel_{chunk_idx+1}_of_{num_chunks}.csv"
                        )
                        chunk_df.drop(columns=['domain']).to_csv(panel_path, index=False)
                        self.log(f"Saved {panel_name} panel {chunk_idx+1}/{num_chunks} "
                                f"({len(chunk_df)} records) to {panel_path}")
            else:
                self.log("No email records available for panelization")

            return True

        except Exception as e:
            self.log(f"Panelization failed: {str(e)}")
            return False




    def step_metrics_reporting(self):
        import traceback  # For detailed error logging
        try:
            manufacturer = self.manufacturer_combo.get().replace(" ", "_")
            manufacturer_title = self.manufacturer_combo.get().title()
            report_path = Path(self.output_entry.get()) / f"{manufacturer}_{self.month_entry.get()}_{self.year_entry.get()}_final_report.md"

            with open(report_path, "w", encoding="utf-8") as f:
                f.write(f"# {manufacturer_title} Data Processing Report\n")
                f.write(f"**Month/Year:** {self.month_entry.get()}/{self.year_entry.get()}\n")
                f.write(f"**Processing Date:** {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write(f"**Sequence #:** {self.sequence_entry.get()}\n\n")

                # Stepwise Rejection Table
                stepwise = self.tracker.get('stepwise', [])
                if stepwise:
                    import pandas as pd  # Ideally, move to top of file
                    f.write("\n## Stepwise Processing Metrics\n")
                    metrics_df = pd.DataFrame(stepwise)
                    f.write(metrics_df.to_markdown(index=False))
                    f.write("\n\n")

                # Original Counts
                f.write("## Original Counts\n")
                f.write(f"- Total Initial Records: {self.tracker.get('initial_count', 0):,}\n")
                f.write(f"- Blank ZIP Codes: {self.tracker.get('blank_zip', 0):,}\n")
                f.write(f"- Blank Names: {self.tracker.get('blank_name', 0):,}\n\n")

                # Processing Rejections
                f.write("## Processing Rejections\n")
                f.write(f"- Invalid Emails: {self.tracker.get('email_rejected', 0):,}\n")
                f.write(f"- Bad Geography (State/ZIP): {self.tracker.get('geo_rejected', 0):,}\n")
                f.write(f"- Business Exclusions: {self.tracker.get('business_rejected', 0):,}\n")
                f.write(f"- Invalid VINs: {self.tracker.get('vin_rejected', 0):,}\n")
                f.write(f"- Blank Sales Type Dropped: {self.tracker.get('blank_sales_type', 0):,}\n")
                f.write(f"- Invalid Sales Type: {self.tracker.get('invalid_sales_type', 0):,}\n\n")

                # Deduplication
                f.write("## Deduplication\n")
                f.write(f"- Total Duplicates Removed: {self.tracker.get('dedup_removed', 0):,}\n\n")

                # Final Output
                f.write("## Final Output\n")
                f.write(f"- Email Records: {self.tracker.get('email_count', 0):,}\n")
                f.write(f"- Mail Records: {self.tracker.get('mail_count', 0):,}\n\n")

                # Domain Breakdown
                f.write("## Email Domain Breakdown\n")
                domain_counts = self.tracker.get('domain_counts', {})
                for domain, count in domain_counts.items():
                    f.write(f"- {domain.title()}: {count:,}\n")

                # Sales Type Breakdown
                f.write("\n## Sales Type Breakdown\n")
                sales_types = self.tracker.get('sales_type_counts', {})
                for stype, count in sales_types.items():
                    f.write(f"- {stype.title()}: {count:,}\n")

                # Sanity Check
                total_removed = (
                    self.tracker.get('blank_name', 0) +
                    self.tracker.get('email_rejected', 0) +
                    self.tracker.get('geo_rejected', 0) +
                    self.tracker.get('business_rejected', 0) +
                    self.tracker.get('vin_rejected', 0) +
                    self.tracker.get('dedup_removed', 0)
                )
                final_count = self.tracker.get('email_count', 0) + self.tracker.get('mail_count', 0)
                start_count = self.tracker.get('initial_count', 0)

                # Blank Name Removal Report
                f.write("\n## Blank Name Removal Report\n")
                f.write(f"- **Records before filtering**: {start_count}\n")
                f.write(f"- **Records after blank name removal**: {start_count - self.tracker.get('blank_name', 0)}\n")
                f.write(f"- **Blank names removed**: {self.tracker.get('blank_name', 0)}\n")
                f.write(f"- [Download removed records](name_removed_blanks.csv)\n\n")

                # Sanity Check Details
                f.write("## Sanity Check Details\n")
                f.write(f"- Start Count: {start_count:,}\n")
                f.write(f"- Total Removed: {total_removed:,}\n")
                f.write(f"- Final Count: {final_count:,}\n")
                f.write(f"- Expected Final: {start_count - total_removed:,}\n")
                f.write(f"- Status: {'✅ Match' if (start_count - total_removed) == final_count else '❌ Mismatch'}\n")

            # Save removed records (with check for 'name' column)
            if hasattr(self, "df") and "name" in self.df.columns:
                blank_name_records = self.df[self.df['name'].isna() | (self.df['name'].str.strip() == '')]
                blank_name_records.to_csv(Path(self.output_entry.get()) / "name_removed_blanks.csv", index=False)

            self.log("Metrics report and blank name removal file generated.")
            return True

        except Exception as e:
            self.log(f"Metrics reporting failed: {str(e)}")
            traceback.print_exc()
            from tkinter import messagebox
            messagebox.showerror(
                "Metrics Reporting Error",
                f"Failed to generate metrics report:\n{str(e)}",
                parent=self.root
            )
            return False




    def export_to_excel_with_formatting(self, df, path):
        """
        Save DataFrame to Excel with header formatting and centered columns.
        Shows error dialogs on failure and logs all actions.
        """
        try:
            df.to_excel(path, index=False)
            wb = load_workbook(path)
            ws = wb.active

            # Format headers: uppercase and bold
            for cell in ws[1]:
                cell.value = str(cell.value).upper()
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Center specific columns if present
            center_cols = ['ROTATION', 'VERSION', 'MOD YR']
            for col_name in center_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')

            wb.save(path)
            self.log(f"Excel file formatted and saved to {path}")
            return True

        except PermissionError:
            msg = f"Permission denied: Could not save {path}. Is the file open in Excel?"
            self.log(msg)
            messagebox.showerror("Save Error", msg, parent=self.root)
            return False

        except Exception as e:
            msg = f"Excel export failed: {str(e)}"
            self.log(msg)
            messagebox.showerror("Export Error", msg, parent=self.root)
            return False








    def generate_summary_report(self):
        """
        Export the pipeline audit trail as a CSV and show a message box.
        """

        try:
            metrics = self.tracker.get_metrics()
            summary = pd.DataFrame(list(metrics.items()), columns=["Metric", "Count"])
            summary_path = Path(self.output_entry.get()) / "pipeline_summary.csv"
            summary_path.parent.mkdir(parents=True, exist_ok=True)
            summary.to_csv(summary_path, index=False)
            self.log(f"Pipeline summary exported to {summary_path}")
            messagebox.showinfo("Summary Exported", f"Pipeline summary exported to:\n{summary_path}", parent=self.root)
            return True
        except Exception as e:
            self.log(f"Failed to generate summary: {str(e)}")
            messagebox.showerror("Summary Error", f"Failed to generate summary:\n{str(e)}", parent=self.root)
            return False



if __name__ == "__main__":
    """Entry point for the application"""
    try:
        root = ThemedTk(theme="aquativo")
        app = DataProcessorApp(root)
        root.mainloop()
    except Exception as e:
            import traceback
            print(f"Critical error: {e}\n{traceback.format_exc()}")
            # Optional: Log to file or show messagebox

